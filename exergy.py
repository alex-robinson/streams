######
## Should be loaded in the pylab environment: ipython -pylab
## Call: execfile("streams.py")

import os,sys,subprocess,string
from collections import OrderedDict
import openpyxl as xl                 # For writing/reading excel files

from copy import deepcopy
import re

def load_reference(filename="ReferenceTables.xlsx"):
    '''
    Load the reference data from the excel tables.
    '''

    # Load the workbook from existing file
    book = xl.load_workbook(filename)
    
    # Get all existing sheetnames
    sheetnames = book.get_sheet_names()
    
    # Load the reference table sheet - the first sheet
    sheet1 = book.get_sheet_by_name("SubstanceRefTables")
    
    ## Now load some reference data

    # First get T0 and p0 used for these data
    T0 = sheet1.cell('B2').value
    p0 = sheet1.cell('A2').value 
    
    # Now get the names of all the substances we
    # have data for (in column C) and the 
    # relevant information
    
    refs = OrderedDict()

    for j in arange(2,1000):
        
        # First see if this row has a substance
        nm = sheet1.cell(row=j,column=2).value

        # If an element was actuall found in this row
        # store all the reference values in the OrderedDict!
        if not nm == None:
            refs[nm] = OrderedDict()
            refs[nm]['T0']      = T0
            refs[nm]['p0']      = p0
            refs[nm]['MW']      = sheet1.cell(row=j,column=3).value
            refs[nm]['e_ch_0a'] = sheet1.cell(row=j,column=4).value    # Ahrends
            refs[nm]['e_ch_0b'] = sheet1.cell(row=j,column=5).value    # Szargut
            refs[nm]['cp_0']    = sheet1.cell(row=j,column=7).value
            refs[nm]['h_0']     = sheet1.cell(row=j,column=8).value
            refs[nm]['s_0']     = sheet1.cell(row=j,column=9).value
            
            refs[nm]['H+']      = sheet1.cell(row=j,column=12).value
            refs[nm]['S+']      = sheet1.cell(row=j,column=13).value
            refs[nm]['a']       = sheet1.cell(row=j,column=14).value
            refs[nm]['b']       = sheet1.cell(row=j,column=15).value
            refs[nm]['c']       = sheet1.cell(row=j,column=16).value
            refs[nm]['d']       = sheet1.cell(row=j,column=17).value
        else:
            break
    
    # Add a default for unknown references
    nm = 'unknown'
    refs[nm] = deepcopy(refs['H2O'])
    refs[nm]['MW'] = 1.0
    refs[nm]['e_ch_0a'] = 0.0
    refs[nm]['e_ch_0b'] = 0.0
    refs[nm]['cp_0'] = 0.0
    refs[nm]['h_0']  = 0.0
    refs[nm]['s_0']  = 0.0
    refs[nm]['H+']   = None
    refs[nm]['S+']   = None
    refs[nm]['a']    = None
    refs[nm]['b']    = None
    refs[nm]['c']    = None
    refs[nm]['d']    = None


    text = "=" * 80
    text = text + "\nThe following substances are available for exergy calculations:\n"
    for i,key in enumerate(refs.keys()):
        text = text + key 
        if not i+1 == len(refs.keys()): text = text + ", "
        if mod(i+1,12) == 0: text = text + "\n"
    text = text + "\n" + "=" * 80
    
    print text
    
    return refs

### THE FUNCTIONS BELOW ARE METHODS
### FOR THE 'ref' object,
### which is an OrderedDict of OrderedDicts
### Eventually these should be encapsulated nicely!!!

def ref_check(refs,nm='N2'):
    '''
    Nicely print the reference values for the desired substance.
    '''
    print "Substance: " + nm

    if nm in refs.keys():
        for key,value in refs[nm].items():
          print "    {:<10} = {}".format(key,value)

    else:
        print "Error: substance not found in reference values!"

    return

def ref_array(refs):
    
    # Get column/row headings
    tmp1,tmp2 = refs.items()[0]
    headings = tmp2.keys()
    names = refs.keys()

    # Determine length of output array
    ns = len(refs)
    nv = len(headings)
    
    # Generate array and fill it in
    table = zeros([ns,nv])
    i = -1
    for nm in names:
        i = i+1
        j = -1
        for key,val in refs[nm].items():
            j = j+1
            table[i,j] = val
    
    return(table)

def ref_print(refs):
    
    nms = refs
    for nm in refs.keys():
      print "Substance: " + nm
      for key,value in refs[nm].items():
          print " {:<10} = {}".format(key,value)
    
    return

### NOW STREAM CLASSES ###

class substance:
    ''' 
    This class manages calculations for a given substance.
    '''

    def __init__(self,id,name,T,p,mdot,x,T0=298.15,p0=1.013):
        
        # Define some constants
        R = 8.314       # kJ/kmol-K, ideal gas constant
        
        # Make a string of the id,name for easy formatting
        idstr = "{:4}".format(id)
        namestr = "{:11}".format(name)
        unknown = False

        # Adjust x values if they are tiny to simplify calcs.
        #if abs(x) < 1e-7: x = 0.0

        # Make sure x is valid
        if x < 0 or x > 1:
            err = '''
Stream {}: {}: Error: molar fraction x must be between 0 and 1!
'''.format(idstr,namestr)
            sys.exit(err)

        # Generate the initial state of this substance in the stream
        state = OrderedDict(T=T,p=p,mdot=mdot*x,x=x)

        # Now check that we can actually model this substance
        nameref = name 
        if not name in refs.keys():
            if x > 0:
                print "Stream {}: {}: Warning: no reference values found for substance! " \
                      "Setting x = {:10.5g} => 0.0.".format(idstr,namestr,x)
            x = 0.0
            nameref = "unknown"
            unknown = True
        
        # Since we know the substance exists,
        # load up the reference values here
        ref = refs[nameref].copy()
        #ref = deepcopy(refs[nameref])  # Shallow copy should be good enough since we don't modify refs!

        # Check to make sure we are using the same reference state (T0,p0)
        if not (T0 == ref['T0'] and p0 == ref['p0']):
            err = '''
Stream {}: {}: Error: reference state is not the same as the reference state of the stream.
                      Stream (T0,p0) = ({:7.3f},{:7.3f})
                   Substance (T0,p0) = ({:7.3f},{:7.3f})
'''.format(idstr,namestr,T0,p0,ref['T0'],ref['p0'])
            sys.exit(err)

        # Get a useful factor (ajr: what is this?)
        y = T/1e3

        ##### Calculations of enthalpy h 
        ##### (depends only on T, because we assume ideal gases)

        # If H+ not known for the element, then we
        # calculate it using cp_0, h_0 and T
        if ref['H+'] == None:
            state['h'] = ( (ref['cp_0']) * (T-ref['T0']) ) + ref['h_0']      # NO change in cp
        
        else:
            state['h'] = 1e3 * (  ref['H+'] + ref['a']*y 
                                + ref['b']/2 * y**2 
                                - ref['c']*y**(-1) 
                                + ref['d']/3 * y**3 )
            
            # Limit enthalpy errors
            if state['h'] > -5.0 and state['h'] < 0.0:
                print "Stream {}: {}: Warning: small negative enthalpy set to zero.".format(idstr,namestr)
                state['h'] = 0.0


        ##### Calculations of entropy
        ##### (depend both on T and p)

        # If H+ not known for the element, then we
        # calculate it using cp_0, h_0 and T and p
        if ref['S+'] == None:
            state['s'] = ( (ref['cp_0']/T)*(T-ref['T0']) - R*log(p/ref['p0']) ) + ref['s_0']
        
        else:
            state['s'] = (  ref['S+'] + ref['a']*log(T) + ref['b']*y 
                          - ref['c']/2 * y**(-2) 
                          + ref['d']/2 * y**2   )     - R * log(p/ref['p0'])

        ##### Calculations of specific heat cp
        ##### (depends on T)
        if ref['a'] == None:
            state['cp'] = ref['cp_0'] 
        
        else:
            state['cp'] = ( ref['a'] + ref['b']*y 
                          + ref['c'] * y**(-2) 
                          + ref['d'] * y**2 )


        # Store the output
        self.id      = id
        self.idstr   = idstr
        self.name    = name
        self.namestr = namestr
        self.unknown = unknown
        self.state   = state
        self.ref     = ref

        return
        
    def __str__(self):
        '''Output the substance object to the screen in a human readable way.'''
        
        text = "\nStream {}: {}".format(self.idstr,self.namestr)
        for key,value in self.state.items():
            text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))
        
        text = text +  "\nREFERENCE VALUES"
        for key,value in self.ref.items():
            text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text 


class stream:
    '''
    This class manages all values concerning a
    given stream and calculation of its exergy.
    '''

    def __init__(self,id,T,p,mdot,composition,phase=[1.0,0.0,0.0],T0=298.15,p0=1.013):
        '''
        Initialize a stream
          id = stream number/name
          T  = stream temperature
          p  = stream pressure
          mdot = stream total mass flow rate kg/hr ?
          composition = list of substances and molar fractions x 
                        ( a list of tuples: [ ('N2',0.5), ('O2',0.1) ] )
          phase = [VFRAC,LFRAC,SFRAC], fraction of vapor, liquid and solid
                  *** phase only needed for H2O streams ***
          T0, p0 = reference state values for temp (K) and pressure (bar)
        '''
        
        ### TO DO ###
        # Implement optional arguments of an array row containing
        # all stream info and an array row containing the header
        # Then extract the info as needed for this class.
        #############
        
        # Make a string of the id for easy formatting
        idstr = "{:4}".format(id)

        # First store the state and the phase
        state = OrderedDict(T=T,p=p,mdot=mdot,T0=T0,p0=p0,phase=phase)

        ## Loop over the substances and initialize each
        ## one for the current stream state
        comp = OrderedDict()
        for name,x in composition:
            comp[name] = substance(id,name,T=T,p=p,mdot=mdot,x=x,T0=T0,p0=p0)
        
        ## Calculate MW of the stream and 
        ## the total molar fraction (should sum to 1 unless substances are missing)
        state['MW'] = 0.0
        xtot        = 0.0
        for key,sub in comp.items():
            state['MW'] = state['MW'] + ( sub.ref['MW']*sub.state['x'] )
            xtot        = xtot + sub.state['x']
        
        # Check if xtot is 1
        # if abs(1.0 - xtot) > 1e-5:
        #     print "Stream {}: Warning: total x not equal to 1!").format(idstr)

        ## Determine what kind of stream we have (H2O or not, flue gas or not)
        isH2O = True 
        for key in comp.keys():
            if not key in ['H2O(l)','H2O(g)']:
                isH2O = False
                break
        
        ## Handling water ##
        ## Make sure that if a stream has H2O(l), it also has H2O(g) and vice-versa
        ## (This will facilitate calculations later involving liquid and gas)
        if 'H2O(l)' in comp.keys() and not 'H2O' in comp.keys():
            comp['H2O'] = substance(id,'H2O',T=T,p=p,mdot=mdot,x=0.0,T0=T0,p0=p0)
        if 'H2O' in comp.keys() and not 'H2O(l)' in comp.keys():
            comp['H2O(l)'] = substance(id,'H2O(l)',T=T,p=p,mdot=mdot,x=0.0,T0=T0,p0=p0)
        
        # Make a duplicate list of stream substances
        # that will be adjusted for calculating the standard (T0,p0) values
        comp0 = deepcopy(comp)
        
        ## If stream is flue gas, adjust water percentages
        ## (For now, always adjust since we don't know how to tell
        ##  if it's flue gas or not)
        # At 25C the pressure would be 0.0317bar, so:
        # x_H2O_new(g) = ( 0.0317 * ( 1- x_H2O ) ) / (1 - 0.0317)         # % mol
        # x_H2O(l) = x_H2O(l) + (x_H2O - x_H2O_new(g))
        if 'H2O(l)' in comp0.keys():
            x_l = comp0['H2O(l)'].state['x']
            x_g = comp0['H2O'].state['x']
            x       = x_l + x_g
            if x > 0.0:
                x_new_g = 0.0317*(xtot-x) / (xtot-0.0317)    # mol frac
                x_new_g = max(x_new_g,0.0)
                x_new_l = (x - x_new_g)                    # mol frac
                x_new_l = max(x_new_l,0.0)
                comp0['H2O(l)'].state['x'] = x_new_l
                comp0['H2O'].state['x']    = x_new_g
                
                text = (
"Stream {}: Check: x of H2O(l),H2O = {:5.3f},{:5.3f} => {:5.3f},{:5.3f}".format(idstr,x_l,x_g,x_new_l,x_new_g) )

                print text

            if not (comp0['H2O(l)'].state['x'] + comp0['H2O'].state['x'] 
                ==  comp['H2O(l)'].state['x'] +  comp['H2O'].state['x']):
                text = "Stream {}: Warning: corrected H2O molar fractions "\
                       "do not sum to original total!".format(idstr)
                print text

        ## Calculate enthalpy (h) and entropy (s) depending on type of stream
        if not isH2O:
            
            # Enthalpy and entropy
            state['h']    = 0.0    # Enthalpy at T for mixture 
            state['h_0']  = 0.0    # Enthalpy at T0 for mixture
            state['s']    = 0.0    # Entropy at T for mixture
            state['s_0']  = 0.0    # Entropy at T0 for mixture
        
            # Loop over elements and get stream variables
            for key,sub in comp.items():
                state['h']   = state['h']   + ( sub.state['h']*sub.state['x'] )
                state['s']   = state['s']   + ( sub.state['s']*sub.state['x'] )
            
            # Loop over standard (T0,p0) elements and get stream variables
            for key,sub in comp0.items(): 
                state['h_0'] = state['h_0'] + ( sub.ref['h_0']*sub.state['x'] )    
                state['s_0'] = state['s_0'] + ( sub.ref['s_0']*sub.state['x'] )             

            
        else:   # Case 2: water stream, H2O(l) and/or H2O(g)
            
            #!!!! CHECK IF WE HAVE SATURATED WATER OR STEAM  !!!!!
            isSaturated = False
            ## ADD TEST HERE

            if isSaturated:
                # Calculate enthalpy from the IAPWS IF97 steam tables for saturated streams
                ## TO BE IMPLEMENTED !!!
                pass

            else:  # Not saturated, calculate here...
                
                # Copy enthalpy and entropy from the IAPWS IF97 steam tables for non-saturated streams
                # h_T_str  = ( (cp_T_el - cp_T0_el) * (T_str-T0) ) + h_T0_el        # here el: H2O
                # s_Tp_str = ( (cp_T_el - cp_T0_el) * ln(T_str/T0) + s_Tp0_el
                pass

            # For now, set values to zero, since we can't calculate anything
            state['h']   = 0.0
            state['s']   = 0.0
            state['h_0'] = 0.0
            state['s_0'] = 0.0
        ####

        self.id    = id
        self.idstr = idstr
        self.state = state
        self.comp  = comp 
        self.comp0 = comp0

        return

    def calc_exergy(self,exergy_type="Ahrends"):
        '''
        Calculation of exergies given the state of 
        a stream.
        '''
        
        # Define some constants
        R = 8.314       # kJ/kmol-K, ideal gas constant
        
        # Get idstr
        idstr = self.idstr

        # Decide which chemical exergy value to use (Ahrends or Szargut)
        if exergy_type == "Ahrends":
            name_ch = 'e_ch_0a'
        elif exergy_type == "Szargut":
            name_ch = 'e_ch_0b'
        else:
            err = '''
Stream {}: Error: Incorrect exergy type given: {}
                    Only "Ahrends" or "Szargut" are allowed.
'''.format(idstr,exergy_type)
            sys.exit(err)


        # Open variables locally
        state = self.state
        comp  = self.comp 

        ## PHYSICAL EXERGY
        # Physical specific exergy of stream  ###
        state['e_ph']    = ( state['h']-state['h_0'] 
                           - state['T0']*(state['s']-state['s_0']) )    # kJ/kmol
        
        # Limit e_ph to postive values and/or give warnings
        if state['e_ph'] < -5.0:
            print "Stream {}: Warning: negative physical exergy.".format(self.id)
        elif state['e_ph'] < 0.0:
            state['e_ph'] = 0.0


        ## CHEMICAL EXERGY
        # Chemical specific exergy of stream ###
        sum1 = 0.0
        sum2 = 0.0
        for key,sub in comp.items():
            if not sub.state['x'] == 0.0:
                sum1 = sum1 + ( sub.ref[name_ch]*sub.state['x'] )
                sum2 = sum2 + ( sub.state['x']*log(sub.state['x']) )

        state['e_ch']    = sum1 + R*state['T0']*sum2           # kJ/kmol
        
        
        ## Convert exergies to per kg
        state['e_ph_kg'] = state['e_ph'] /state['MW']          # kJ/kg
        state['e_ch_kg'] = state['e_ch'] /state['MW']          # kJ/kg
        
        ## Get absolute exergies (in MW)
        state['E_ph'] = state['e_ph_kg'] *state['mdot']/1e3    # MW
        state['E_ch'] = state['e_ch_kg'] *state['mdot']/1e3    # MW

        ## TOTAL EXERGY

        # Specific
        state['e_tot_kg'] = state['e_ph_kg'] + state['e_ch_kg']   # KJ/kg
        
        # Absolute
        state['E_tot'] = state['E_ph'] + state['E_ch']            # MW
        
        E_tot_check = state['e_tot_kg'] * state['mdot'] /1e3      # MW
        
        ebs = 1e-5
        if not abs(state['E_tot']-E_tot_check) < ebs:
            err = '''
Stream {}: Warning: total exergy calculations do not match!
                  E_ph + E_ch != e_tot_kg * mdot
                  {:10.3f} + {:10.3f} != {:10.3f} * {:10.3f}
'''.format(idstr,state['E_ph'],state['E_ch'],state['e_tot_kg'],state['mdot']/1e3)
            

            #sys.exit(err)
            print err
        
        # Update the stream's state object
        self.state = state 

        return
    
    def __repr__(self):
        
        text = "Stream: {}".format(self.id)
         
        for key,value in self.state.items():
            if key == "phase":
                text = text + "\n {:<12} = {:>5.2n},{:>5.2n},{:>5.2n}".format(key,float(value[0]),float(value[1]),float(value[2]))
            else:
                text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text

    def __str__(self):
        '''Output the stream object to the screen in a human readable way.'''
        
        text = "Stream: {}".format(idstr)
         
        for key,value in self.state.items():
            if key == "phase":
                text = text + "\n {:<12} = {:>5.2n},{:>5.2n},{:>5.2n}".format(key,float(value[0]),float(value[1]),float(value[2]))
            else:
                text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        # text = text +  "\n====== reference values ======"
        # for key,value in self.ref.items():
        #     text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text 


### SIMULATION CLASS ###
class simulation:
    '''This class holds all simulation information in the
       proper format. This can be read in from an excel file 
       or original formats from aspen (not implemented) or ebsilon (not implemented)
       The simulation data is held in stream classes.
    '''
    
    def __init__(self,filename="ExampleSimulation.xlsx",
                 sheetname="ExampleStreams1",exergy_method=None):
        '''Initialize the simulation'''
        
        ## Check file extension to see which type of data to load
        ext = filename.rsplit(".",1)[1]
        if ext == "xlsx":  # load excel
            self.streams = self.load_excel(filename=filename,sheetname=sheetname)
            newfilename = filename
        elif ext == "rep": # load aspen
            print filename
            self.streams = self.load_aspen(filename=filename)
            #newfilename = filename0.rsplit(".",1)[0] + ".xlsx"
        else:
            pass # Incorrect format or unimplemented format

        ## Now we have the streams object

        # Output some information about the streams
        

        ## Calculate exergy now
        if exergy_method == "gatex":
            pass
        elif exergy_method == "default":
            
            for key,stream in self.streams.items():
                stream.calc_exergy()



        return 
    
    def load_excel(self,filename="ExampleSimulation.xlsx",sheetname="ExampleStreams1"):
        '''Load the simulation data from an excel sheet.'''

        # Load the workbook from existing file
        book = xl.load_workbook(filename)
        
        # Get all existing sheetnames
        sheetnames = book.get_sheet_names()
        
        # Make sure the sheet exists!
        if not sheetname in sheetnames:
            err = '''
Error:: load_excel:: Desired sheetname {} does not exist in the workbook {}.
'''.format(filename,sheetname)
            sys.exit(err)

        # Load the simulation sheet
        sheet1 = book.get_sheet_by_name(sheetname)
        
        #### Begin loading simulation data

        streams = OrderedDict()

        for j in arange(1,1000):    # Max 1000 streams!
            
            # First see if this row has a substance
            streamid = sheet1.cell(row=j,column=0).value

            # If an element was actually found in this row,
            # load all the data and generate a new stream object
            if not streamid == None:
                T         = sheet1.cell(row=j,column=1).value
                p         = sheet1.cell(row=j,column=2).value
                mdot      = sheet1.cell(row=j,column=3).value
                # MW is in column 4, but is calculated internally
                vfrac     = sheet1.cell(row=j,column=5).value
                lfrac     = sheet1.cell(row=j,column=6).value
                sfrac     = sheet1.cell(row=j,column=7).value
                
                if vfrac == None: vfrac = 0.0
                if lfrac == None: lfrac = 0.0
                if sfrac == None: sfrac = 0.0

                # Now loop over substances
                comp = []
                for i in arange(0,500):   # Max 500 substances!
                    
                    name = sheet1.cell(row=0,column=8+i).value
                    
                    if not name == None:
                        x = sheet1.cell(row=j,column=8+i).value
                        if x == None: x = 0
                        #print i,j,name,x
                        comp.append( (name,x) )
                    else:
                        break

                # Now we have loaded the state variables and 
                # the composition, so generate a stream object
                streamidstr = "{}".format(streamid)
                streams[streamidstr] = stream(id=streamid,T=T,p=p,mdot=mdot,
                                              phase=[vfrac,lfrac,sfrac],composition=comp)

            else:
                break
        
        return streams
    
    def write_excel(self,filename,sheetname,writeSim=True,writeResults=True):
        '''Write simulation results to an excel file.'''
        
        # Load streams locally
        streams = self.streams

        # Generate a sheetname for the results
        sheetname_results    = sheetname + "_exergy"
        sheetname_substances = sheetname + "_sub"

        # Check if file already exists
        isFile = os.path.isfile(filename)

        if isFile:
            # Open existing workbook
            book = xl.load_workbook(filename)
            
            # Check if our desired worksheet already exists
            # and adjust sheetname as needed.
            sheetnames = book.get_sheet_names()
            if sheetname in sheetnames:
                print "{} already exists in {}.".format(sheetname,filename)
                choice = raw_input("Overwrite these results (y/n)? ")
                if not choice in ["y","Y","yes","YES","Yes"]:
                    sheetname = sheetname + "_new"

            # If sheetname already exists, delete it so we can start fresh.
            if sheetname in sheetnames:
                sheet1 = book.get_sheet_by_name(sheetname)
                book.remove_sheet(sheet1)
            
            # Now generate a fresh sheet with the right name
            sheet1 = book.create_sheet()        # Create a new sheet
            sheet1.title = sheetname            # Make sure the new sheet has the right name
            
            # If sheetname already exists, delete it so we can start fresh.
            if sheetname_results in sheetnames:
                sheet2 = book.get_sheet_by_name(sheetname_results)
                book.remove_sheet(sheet2)
            
            # Make a second fresh sheet for the results
            sheet2 = book.create_sheet()
            sheet2.title = sheetname_results
            
            # If sheetname already exists, delete it so we can start fresh.
            if sheetname_substances in sheetnames:
                sheet3 = book.get_sheet_by_name(sheetname_substances)
                book.remove_sheet(sheet3)
            
            # Make a second fresh sheet for the results
            sheet3 = book.create_sheet()
            sheet3.title = sheetname_substances

        
        else:
            # Generate a new workbook
            book = xl.Workbook()
            
            # Get the first sheet and rename it with the right name.
            sheet1 = book.get_active_sheet()    # Active sheet in new book is always the first one
            sheet1.title = sheetname            # Make sure the first sheet has the right name
            
            # Make a second sheet for the results
            sheet2 = book.create_sheet()
            sheet2.title = sheetnameresults
        
        ## Now we should have an open workbook
        ## with one or two empty sheets
        
        ## First write the simulation data to sheet1
        
        ## Make a header
        header1 = ["T","p","mdot","MW","VFRAC","LFRAC","SFRAC"]
        key = streams.keys()[0]
        header2 = streams[key].comp.keys()
        header = header1 + header2

        # Define the column offset for the headers
        offset  = 1 
        offset2 = 8

        sheet1.cell(row=0,column=0).value = "Stream"
        for j,head in enumerate(header):
            sheet1.cell(row=0,column=offset+j).value = head 
        
        # Loop over the streams and write each line of data to the file
        for i,key in enumerate(streams.keys()):
              
              state = streams[key].state

              # Write the stream id and state variables
              sheet1.cell(row=i+1,column=offset-1).value = key
              sheet1.cell(row=i+1,column=offset+0).value = state['T']
              sheet1.cell(row=i+1,column=offset+1).value = state['p']
              sheet1.cell(row=i+1,column=offset+2).value = state['mdot']
              sheet1.cell(row=i+1,column=offset+3).value = state['MW']
              sheet1.cell(row=i+1,column=offset+4).value = state['phase'][0]
              sheet1.cell(row=i+1,column=offset+5).value = state['phase'][1]
              sheet1.cell(row=i+1,column=offset+6).value = state['phase'][2]

              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header2):
                  if vnm in streams[key].comp.keys():
                      val = streams[key].comp[vnm]
                      sheet1.cell(row=i+1,column=offset2+j).value = val.state['x']

        ## Now write the results to sheet2

        ## Make a header
        header = ["T","p","mdot","MW","h","s","h_0","s_0","e_ph","e_ch","e_ph_kg","e_ch_kg","e_tot_kg","E_ph","E_ch","E_tot"]
        
        # Define the column offset for the header
        offset = 1 
        
        sheet2.cell(row=0,column=0).value = "Stream"
        for j,head in enumerate(header):
            sheet2.cell(row=0,column=offset+j).value = head 
        
        # Loop over the streams and write each line of data to the file
        for i,key in enumerate(streams.keys()):
              
              # Write the stream id
              sheet2.cell(row=i+1,column=offset-1).value = key
              
              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header):
                  if vnm in streams[key].state.keys():
                      val = streams[key].state[vnm]
                      sheet2.cell(row=i+1,column=offset+j).value = val
        
        ## Now write the individual substance information in sheet3
        
        ## Make a header
        header1 = ["T","p","mdot","x","h","s","cp"]
        header2 = ["MW","T0","p0","e_ch_0a","e_ch_0b","cp_0","h_0","s_0","H+","S+","a","b","c","d"]
        header3 = ["x_25","h_0_25","s_0_25"]

        header  = header1 + header2 + header3

        # Define the column offset for the header
        offset  = 1 
        offset2 = len(header1) + offset
        offset3 = len(header1) + len(header2) + offset

        # Start a row counter
        row = 0 

        # Loop over the streams and write each line of data to the file
        for i,key0 in enumerate(streams.keys()):
            
            row = row + 1

            # Write the header again (for each stream)
            sheet3.cell(row=row,column=0).value = "Stream"
            for j,head in enumerate(header):
                sheet3.cell(row=row,column=offset+j).value = head 
        
                
            # Get the current stream
            stream = streams[key0]

            ## Now loop over each substance in the stream
            for j,key in enumerate(stream.comp):
              
              state  = stream.comp[key].state
              ref    = stream.comp[key].ref
              state0 = stream.comp0[key].state
              ref0   = stream.comp0[key].ref

              # Increment the row
              row = row + 1
              
              # Write the stream id
              sheet3.cell(row=row,column=offset-1).value = key0
                
              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header1):
                  val = state[vnm]
                  sheet3.cell(row=row,column=offset+j).value = val
              
              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header2):
                  val = ref[vnm]
                  sheet3.cell(row=row,column=offset2+j).value = val
              
              # Output last ones manually since header doesn't match var names
              sheet3.cell(row=row,column=offset3+0).value = state0['x']
              sheet3.cell(row=row,column=offset3+1).value = ref0['h_0']
              sheet3.cell(row=row,column=offset3+2).value = ref0['s_0']

        ## Finally, save the book to the actual excel file
        book.save(filename)

        return 
    
    def load_aspen(self,filename="simu1.rep"):
        '''
        Given the aspen output file located at fldr/file_in,
        convert the format to the stream format 
        '''

        print "Read input from Aspen file: {}".format(filename)

        # Load lines of input file
        lines = open(filename,'r').readlines()
        
        # STEP 1: Filter just to all lines related to stream section

        # Names of subsections
        subsections = ["COMPONENTS: KMOL/HR","TOTAL FLOW:","STATE VARIABLES:"]
        nss = len(subsections)
        
        # Line substrings to skip entirely
        skipstrings = ["(CONTINUED)"]

        in_stream      = False
        in_subsection  = False
        got_subsection = [False] * nss
        skipnext       = False
        sectstart      = False
        
        # Define assumed units
        units_T = "C"
        units_P = "Bar"

        lines1 = []

        # Loop over the lines and add those inside of the 'stream section'
        # Modify some strings to make a proper table
        for line in lines:
            
            # First strip extra whitespace
            line = line.strip()
            addthisline = False 

            ## Skip blank and other unwanted lines
            if line == "" or any([substring in line for substring in skipstrings]): continue 

            ## Determine whether we are in the stream section
            if line == "STREAM SECTION":
                in_stream = True
                continue 

            ## ... or not in the stream section anymore
            if "ASPEN PLUS" in line:
                in_stream = False
                continue
            
            # How many subsections have we looped over so far?
            ngss = sum(got_subsection)
            

            ## Determine if we're not in a subsection we want anymore
            if in_stream and in_subsection:
                if ":" in line: in_subsection = False 

            ## And determine if now we found a subsection we want
            if in_stream and ngss < nss:
                if line == subsections[ngss]:
                    in_subsection        = True 
                    got_subsection[ngss] = True
                    continue     # we don't actually want this line...
            
            ##
            ## Now that we know where we are (in_stream, in_subsection),
            ## Decide what to do with lines
            ##

            ## If we find a new stream heading, add this line to the list
            if in_stream and not in_subsection:
                
                if sectstart and "STREAM ID" in line:

                    # Make sure previous stream set was not empty
                    if len(lines1) > 1:
                        if "STREAM" in lines1[len(lines1)-1]: lines1.pop()
                    
                    addthisline = True 
                    sectstart = False
                
                # If the line only contains the character '-' it's a 
                # section header...
                if all([c in "-" for c in line]):
                    sectstart = True
                    got_subsection = [False] *nss  # reset subsections

            ## ... Else if we're inside a subsection we want,
            ## then we'll add this line to the list
            elif in_stream and in_subsection: 
                addthisline = True
            
            ## If we indeed found a line we want, then
            ## modify the strings and add it to the list!
            if addthisline:
                
                # Make some handy string replacements here
                line1 = line

                # Get some info about units if available
                if "TEMP   C"   in line1: units_T = "C"
                if "TEMP   K"   in line1: units_T = "K"
                if "PRES   BAR" in line1: units_p = "Bar"

                if "STREAM" in line1:
                    line1 = line1.replace("STREAM ID","STREAM   ")
                else:
                    
                    # Remove extra units of T and p
                    line1 = line1.replace(" BAR ","     ")
                    line1 = line1.replace(" C ","   ")
                    
                    # Change some headings
                    line1 = line1.replace("TEMP   C",  "T       ")   # Temperature
                    line1 = line1.replace("PRES   BAR","p         ") # Pressure
                    line1 = line1.replace("KG/HR","mdot ")           # Mass flow rate
                    line1 = line1.replace("KCAL/KG","H      ")       # Enthalpy
                    line1 = line1.replace("CAL/GM-K","SE      ")     # Entropy
                    
                    ### Number corrections

                    # Replace missing numbers with nan
                    line1 = line1.replace("MISSING","nan")

                    # (exponents need the 'e', 1.0-05 => 1.0e-05)
                    matches = re.findall(r'[0-9]\-[0-9]',line1)
                    for substr in matches:
                        substrnew = substr[0:1] + "e-" + substr[2:3]
                        line1 = line1.replace(substr,substrnew)
                    matches = re.findall(r'[0-9]\+[0-9]',line1)
                    for substr in matches:
                        substrnew = substr[0:1] + "e" + substr[2:3]
                        line1 = line1.replace(substr,substrnew)


                # Now add this line to the subset of lines we want
                lines1.append(line1)
                #print line1

        # Remove last line if it belongs to a spurious heading...
        if "STREAM" in lines1[len(lines1)-1]:
            lines1.pop()

        ##########################
        
        # Split the lines into separate lists so that
        # they can properly be merged into an array
        tmplists = []
        q = -1
        for line in lines1:
            if "STREAM" in line:
                tmplists.append([])
                q = q + 1
            tmplists[q].append(line)
        
        # Check that all lists have the same length!!
        print "If these are all the same number, then data was properly read:"
        for q in arange(len(tmplists)):
            print "Length of each page:",len(tmplists[q])
        
        # Save how man pages and how many columns (right now, rows)
        # of data are on each page.
        npages = len(tmplists)
        ncols  = len(tmplists[0])
        
        # Now, generate a new table, by pasting the multiples
        # on as additional columns. In this way, each column
        # represents one stream 
        lines2 = tmplists[0]
        for q in arange(1,npages):
            for i in arange(0,ncols):
                lines2[i] = lines2[i] + "  " + tmplists[q][i]

        # Now split the lines by white space
        # Then loop through and remove columns 
        # where the stream number is actually a string
        tmp = [line.split() for line in lines2]
        #headings = [val[0] for val in tmp]
        
        # Replace stream numbers that are strings with negative number
        for i in arange(len(tmp[0])):
            if not tmp[0][i] == "STREAM":
                try:
                    check = float(tmp[0][i])
                except:
                    tmp[0][i] = "-9999"

        # Split the lines by white space and convert list into
        # a numpy array. Transpose so that each row is one stream
        table0 = asarray( tmp )
        table0 = transpose(table0)

        # Extract the headings of each column and remove them from the data part
        headings = table0[0,:]
        
        # Modify some element names
        # Define aliases: (Aspen_name,Reference_name)
        aliases = [ ("WATER",  "H2O"),
                    ("H2O(g)", "H2O"),
                    (" AR ",     " Ar "),
                    (" S ",      " S(s) "),
                    (" C ",      " C(s) ") ]

        tmp =  "  "+"  ".join(headings)+"  "
        for alias in aliases:
            tmp = tmp.replace(alias[0],alias[1])
        headings = array(tmp.split()) 
        #print "  ".join(headings)

        inds = [ e in ["STREAM","-9999"] for e in table0[:,0] ]        
        table1 = table0[logical_not(inds),:]

        # Convert the text table into a data array
        # and order the rows according to the stream number
        data = asarray( [[float(a) for a in row] for row in table1 ] )
        order = data[:,0].argsort()
        data = data[order,:]
        
        ### SO NOW ABOVE WE HAVE ALL THE SIMULATION DATA IN A TABLE
        ### Now, save it in the streams format
        
        # Indices of substances
        inds = [ not e in ['KMOL/HR','mdot','CUM/HR','TEMP','PRES','VFRAC','LFRAC','SFRAC']
                    for e in headings ]
        ii1 = find(inds)

        # Indices of other information
        inds = [ e in ['KMOL/HR','mdot','CUM/HR','TEMP','PRES','VFRAC','LFRAC','SFRAC']
                    for e in headings ]
        ii2 = find(inds)

        # Set all nan values to zero
        #data[isnan(data)] = 0

        # Unit converters
        conv_T = 0.0
        if units_T == "C": conv_T = 273.15
        conv_mdot = 1/3600.0

        ## Loop over each stream and load the data
        streams = OrderedDict()
        for j,row in enumerate(data):
            
            # Get state variables and id
            streamid = int(row[ii1[0]])
            tot      = row[ii2[0]]
            mdot     = row[ii2[1]] * conv_mdot # KG/HR => KG/S
            T        = row[ii2[3]] + conv_T    # Must be in Kelvin
            p        = row[ii2[4]]
            vfrac    = row[ii2[5]]
            lfrac    = row[ii2[6]]
            sfrac    = row[ii2[7]]
            
            # Check for reasonable temperature values.
            print "Stream {}: T,p,mdot = {},{},{}".format(streamid,T-273.15,p,mdot)

            # Now loop over substances
            comp = []
            for i in ii1[1:]:   # All of ii1 indices, except the first that was the streamid
                
                name = headings[i]
                x    = row[i]
                if tot > 0.0: x = x / tot
                comp.append( (name,x) )

            # Store the variables inside a new stream object
            streamidstr = "{}".format(streamid)
            streams[streamidstr] = stream(id=streamid,T=T,p=p,mdot=mdot,
                                          phase=[vfrac,lfrac,sfrac],composition=comp)
            
        # Add to simulation object
        #self.streams = streams

        return streams
    
    def sim_to_gatex():
        '''Translate a simulation object into an array readable by GATEX.'''

        ## CODE below to output in format for GATEX!!
        ## ajr: need to modify this so that we generate it from the streams object, not the data table.
        
        # Get the shape of the data
        nstreams = data.shape[0]
        ncols    = data.shape[1]
        
        # Make the output array and headings we want for each column
        outh = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])
        ncolsx = len(outh) 
        out = zeros((nstreams,ncolsx))

        # Populate a new array that is exactly the size and shape we need
        # by matching the columns of input data with the right columns in the output arra
        for j in arange(0,ncols):
            h0 = headings[j]
            
            if h0 in outh:
                #print "Writing " + h0 + " to column " + str(j)
                out[:,outh==h0] = data[:,headings==h0]

        ## Make unit conversions
        out[:,outh=="mdot"] = out[:,outh=="mdot"] * (1.0/3600.0)   # kg/h => kg/s
        out[:,outh=="SE"] = out[:,outh=="SE"] * (1e3)         # CAL/GM-K => CAL/kg-K

        # Now convert the output composition values to fractions
        elements = array(["Ar","CO2","CO","COS","H2O","CH4","H2","H2S","N2","O2","SO2"])
        inds = in1d(outh,elements)

        for s in arange(0,nstreams):

            # First eliminate nan values (fortran program can't read them)
            out[s,isnan(out[s,:])] = 0.0
            
            # Now scale mass flows into fractions
            tot = data[s,headings=="KMOL/HR"]
            if tot > 0.0: out[s,inds] = out[s,inds] / tot
            #print "tot, sum:",tot,sum(out[s,inds])  # check each row sums to 1!
        
        # Reorder the data by stream number
        ### Sort the list by row (using the first column to decide the order) ##
        order = out[:,0].argsort()
        out = out[order,:]
        
        # Check that stream numbers go up sequentially,
        # if not, remove the remaining streams
        nsdiff = diff(out[:,0])
        inds = [not e == 1 for e in nsdiff]
        
        if any(inds):
            err ='''
ERROR:: load_aspen:: Stream numbering error. "
  Each stream number should increase sequentially by 1."
  The following stream numbers appear to be out of sequence:"
'''
            for q in arange(len(out[:,0])-1):
                if inds[q]: print "    Stream ",int(out[q+1,0])

            #sys.exit("Stream numbers (see above).")

        ## FINAL STEP: if desired, also write the data to output file in ebsilon format
        if write_mfile: write_ebsilon(out,file_output)
        
        # Return the array of stream data in ebsilon format
        return out 

    def __str__(self):
        '''Print out the handy simulation.'''
        
        text = ""

        for key,stream in self.streams.items():
            text = text + "\n" + repr(stream)


        return text


