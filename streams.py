######
## Should be loaded in the pylab environment: ipython -pylab
## Call: execfile("streams.py")

import os,sys,subprocess,string
from collections import OrderedDict
import openpyxl as xl                 # For writing/reading excel files

## !!! THE STREAM WITH THE REFERENCE PRESSURE AND TEMPERATURE IS STREAM NUMBER 1 !!! ##
  
#-------------------------------------------------------------------##

def load_streams(fldr="./",file_in="CombinedRes1.m",filetype=None):
    '''
    Generic function to load streams. This function
    will determine automatically whether to use
    load_ebsilon or load_aspen based on the input
    filename. Or this can be set with the argument
    filetype="ebsilon"
    '''
    
    if filetype is None:
        filetype = "ebsilon"
        if ".rep" in file_in: filetype = "aspen"

    if filetype == "ebsilon":
        streams = load_ebsilon(fldr,file_in)
    elif filetype == "aspen":
        streams = load_aspen(fldr,file_in)
    else:
        print("Error: file type not recognized: " + filetype)
        streams = None
    
    print("Stream data loaded from: " + os.path.join(fldr,file_in))

    return streams

def load_ebsilon(fldr="./",file_in="CombinedRes1.m"):
    '''
    Load an ebsilon output array of stream data.
    Store data in an 'nparray'.
    If the file doesn't load, exit
    '''
    
    ### Try loading the values into an 'nparray' (numpy array) from the file #####
    ### If the file does not exist and an IOError occurs, exit               #####
    
    # Generate necessary filename
    file_input = fldr + "/" + file_in 

    try:
        out = loadtxt(file_input, skiprows = (1),comments = ']')
    except IOError as e:
        print('Unable to locate the file:', f_in, 'Ending Program.''\n', e)
        input("\nPress any key to exit.")
        sys.exit()
    
    ### Sort the list by row (using the first column to decide the order) ##
    order = out[:,0].argsort()
    out = out[order,:]

    return out

def write_ebsilon(out,file_output):
    '''
    Write an array of stream data to the ebsilon .m output format.
    '''
    
    nstreams = out.shape[0]

    outf = open(file_output,'w')
    outf.write("data{1} = [\n")
    for s in arange(0,nstreams):
        line = ["{0:12.6f}".format(v) for v in out[s,:]]
        line[0] = "{0:3g}".format(out[s,0])   # Correct the stream number to be an integer!
        line = " ".join(line)+"\n"
        outf.write(line)
    outf.write("];\n\n")
    
    print("Stream data written to: " + file_output)

    return 

def load_aspen(fldr="./",file_in="simu1.rep",write_mfile=True):
    '''
    Given the aspen output file located at fldr/file_in,
    convert the format to the .m table for gatex at fldr/out 
    '''

    # Generate necessary filenames
    file_input   = fldr + "/" + file_in     
    file_output = file_input.replace(".rep",".m")
    
    print("Aspen files")
    print("In  :" + file_input)
    print("Out :" + file_output)

    # Load lines of input file
    lines = open(file_input,'r').readlines()
    
    # STEP 1: Filter just to all lines related to stream section

    # Names of subsections
    subsections = ["COMPONENTS: KMOL/HR","TOTAL FLOW:","STATE VARIABLES:"]
    nss = len(subsections)
    
    # Line substrings to skip entirely
    skipstrings = ["(CONTINUED)"]

    in_stream      = False
    in_subsection  = False
    got_subsection = [False] * nss
    skipnext       = False
    sectstart      = False
    
    lines1 = []

    # Loop over the lines and add those inside of the 'stream section'
    # Modify some strings to make a proper table
    for line in lines:
        
        # First strip extra whitespace
        line = line.strip()
        addthisline = False 

        ## Skip blank and other unwanted lines
        if line == "" or any([substring in line for substring in skipstrings]): continue 

        ## Determine whether we are in the stream section
        if line == "STREAM SECTION":
            in_stream = True
            continue 

        ## ... or not in the stream section anymore
        if "ASPEN PLUS" in line:
            in_stream = False
            continue
        
        # How many subsections have we looped over so far?
        ngss = sum(got_subsection)
        

        ## Determine if we're not in a subsection we want anymore
        if in_stream and in_subsection:
            if ":" in line: in_subsection = False 

        ## And determine if now we found a subsection we want
        if in_stream and ngss < nss:
            if line == subsections[ngss]:
                in_subsection        = True 
                got_subsection[ngss] = True
                continue     # we don't actually want this line...
        
        ##
        ## Now that we know where we are (in_stream, in_subsection),
        ## Decide what to do with lines
        ##

        ## If we find a new stream heading, add this line to the list
        if in_stream and not in_subsection:
            
            if sectstart and "STREAM ID" in line:

                # Make sure previous stream set was not empty
                if len(lines1) > 1:
                    if "STREAM" in lines1[len(lines1)-1]: lines1.pop()
                
                addthisline = True 
                sectstart = False
            
            # If the line only contains the character '-' it's a 
            # section header...
            if all([c in "-" for c in line]):
                sectstart = True
                got_subsection = [False] *nss  # reset subsections

        ## ... Else if we're inside a subsection we want,
        ## then we'll add this line to the list
        elif in_stream and in_subsection: 
            addthisline = True
        
        ## If we indeed found a line we want, then
        ## modify the strings and add it to the list!
        if addthisline:

            # Make some handy string replacements here
            line1 = line

            if "STREAM" in line1:
                line1 = line1.replace("STREAM ID","STREAM   ")
            else:
                line1 = line1.replace(" BAR ","     ")
                line1 = line1.replace(" C ","   ")
                line1 = line1.replace("MISSING","nan")
                line1 = line1.replace("+","e")
                line1 = line1.replace("-","e-")
                line1 = line1.replace(" e-"," -")
                line1 = line1.replace("GMe-","GM-")
                
                line1 = line1.replace("TEMP","T   ")
                line1 = line1.replace("PRES","p   ")
                line1 = line1.replace("VFRAC","x    ")
                line1 = line1.replace("WATER","H2O  ")
                line1 = line1.replace("AR","Ar")
                line1 = line1.replace("KG/HR","mdot ")          # Mass flow rate
                line1 = line1.replace("KCAL/KG","H      ")      # Enthalpy
                line1 = line1.replace("CAL/GM-K","SE      ")    # Entropy

            # Now add this line to the subset of lines we want
            lines1.append(line1)
    
    # Remove last line if it belongs to a spurious heading...
    if "STREAM" in lines1[len(lines1)-1]:
        lines1.pop()

##########################
    
    # Split the lines into separate lists so that
    # they can properly be merged into an array
    tmplists = []
    q = -1
    for line in lines1:
        if "STREAM" in line:
            tmplists.append([])
            q = q + 1
        tmplists[q].append(line)
    
    # Check that all lists have the same length!!
    print("If these are all the same number, then data was properly read:")
    for q in arange(len(tmplists)):
        print("Length of each page:",len(tmplists[q]))
    
    # Save how man pages and how many columns (right now, rows)
    # of data are on each page.
    npages = len(tmplists)
    ncols  = len(tmplists[0])
    
    # Now, generate a new table, by pasting the multiples
    # on as additional columns. In this way, each column
    # represents one stream 
    lines2 = tmplists[0]
    for q in arange(1,npages):
        for i in arange(0,ncols):
            lines2[i] = lines2[i] + "  " + tmplists[q][i]

    # Now split the lines by white space
    # Then loop through and remove columns 
    # where the stream number is actually a string
    tmp = [line.split() for line in lines2]
    #headings = [val[0] for val in tmp]
    
    # Replace stream numbers that are strings with negative number
    for i in arange(len(tmp[0])):
        if not tmp[0][i] == "STREAM":
            try:
                check = float(tmp[0][i])
            except:
                tmp[0][i] = "-9999"

    # Split the lines by white space and convert list into
    # a numpy array. Transpose so that each row is one stream
    table0 = asarray( tmp )
    table0 = transpose(table0)

    # Extract the headings of each column and remove them from the data part
    headings = table0[0,:]
    #inds = [ e == "STREAM" for e in table0[:,0] ]
    inds = [ e in ["STREAM","-9999"] for e in table0[:,0] ]
    
    print(headings)
    
    table1 = table0[logical_not(inds),:]

    # Convert the text table into a data array
    data = asarray( [[float(a) for a in row] for row in table1 ] )
    
    # Get the shape of the data
    nstreams = data.shape[0]
    ncols    = data.shape[1]

    # Make the output array and headings we want for each column
    outh = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])
    ncolsx = len(outh) 
    out = zeros((nstreams,ncolsx))

    # Populate a new array that is exactly the size and shape we need
    # by matching the columns of input data with the right columns in the output arra
    for j in arange(0,ncols):
        h0 = headings[j]
        
        if h0 in outh:
            #print("Writing " + h0 + " to column " + str(j))
            out[:,outh==h0] = data[:,headings==h0]

    ## Make unit conversions
    out[:,outh=="mdot"] = out[:,outh=="mdot"] * (1.0/3600.0)   # kg/h => kg/s
    out[:,outh=="SE"] = out[:,outh=="SE"] * (1e3)         # CAL/GM-K => CAL/kg-K

    # Now convert the output composition values to fractions
    elements = array(["Ar","CO2","CO","COS","H2O","CH4","H2","H2S","N2","O2","SO2"])
    inds = in1d(outh,elements)

    for s in arange(0,nstreams):

        # First eliminate nan values (fortran program can't read them)
        out[s,isnan(out[s,:])] = 0.0
        
        # Now scale mass flows into fractions
        tot = data[s,headings=="KMOL/HR"]
        if tot > 0.0: out[s,inds] = out[s,inds] / tot
        #print("tot, sum:",tot,sum(out[s,inds]))  # check each row sums to 1!
    
    # Reorder the data by stream number
    ### Sort the list by row (using the first column to decide the order) ##
    order = out[:,0].argsort()
    out = out[order,:]
    
    # Check that stream numbers go up sequentially,
    # if not, remove the remaining streams
    nsdiff = diff(out[:,0])
    inds = [not e == 1 for e in nsdiff]
    
    if any(inds):
        print("ERROR:: load_aspen:: Stream numbering error. ")
        print("  Each stream number should increase sequentially by 1.")
        print("  The following stream numbers appear to be out of sequence:")
        
        for q in arange(len(out[:,0])-1):
            if inds[q]: print("    Stream ",int(out[q+1,0]))

        #sys.exit("Stream numbers (see above).")

    ## FINAL STEP: if desired, also write the data to output file in ebsilon format
    if write_mfile: write_ebsilon(out,file_output)
    
    # Return the array of stream data in ebsilon format
    return out 

def calc_exergy_gatex(streams,fldr="./",gatex_exec="./gatex_pc_if97_mj.exe",sat_water=[-1.0],sat_steam=[-1.0]):
    '''
    Use this subroutine to calculate exergy from a set of stream data (array)
    using GATEX.

    == INPUT ==
    streams : numpy array of dimensions n_streams X n_variables
    fldr    : folder containing gatex program and 
              where input and output data are stored
              (for now must be the current working directory, since gatex.exe looks there)
    
    == OUTPUT ==
    E       : table of exergy calculations obtained from GATEX 

    '''
    
    print('======================')
    print("Generating GATEX input files")

    # Define important gatex filenames
    gatex_f1 = os.path.join(fldr,"gate.inp")
    gatex_f2 = os.path.join(fldr,"flows.prn")
    gatex_f3 = os.path.join(fldr,"composition.prn")

    # To help figure things out, create a header
    # that shows the variable of each column
    head = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])

    # How many streams and variables are we working with?
    n_streams = len(streams)
    n_var     = len(head)

    ### Create the gate.inp file with the reference values #################
    ref = open(gatex_f1,'w',buffering=0)                        
    t0 = float(streams[0,2])
    p0 = float(streams[0,3])
    ref.write('\n\n[system]\n\nt0 =\t')             
    ref.write(str(t0))
    ref.write(' ;\np0 =\t')
    ref.write(str(p0))
    ref.write(' ;\nnumber =\t')
    ref.write(str(n_streams))
    ref.write('. ;\nndiff =\t')
    ref.write(str(n_streams))
    ref.write('. ;\nkelvin =\t')
    ref.write('0. ;')  
    ref.close()
    print(gatex_f1 + " is closed? " + str(ref.closed))
    print("Wrote file for GATEX: " + gatex_f1)

    #----------------------------------------------------------------------#

    #-- Separate streams depending on their type --------------------------#
    for i in arange(0,n_streams):

        ## NEW (30.10.2012): when no CH4, H2, N2 or O2 are present, then stream is H2O
        if streams[i,12] + streams[i,13] + streams[i,15] + streams[i,16] < 0.0000001:
            streams[i,10] = 1.0

        ## END NEW ##
        
        if streams[i,10] == 1.0:            # H2O stream

            if streams[i,4] > 0.001 and streams[i,4] < 0.999:
                pass  # Take original phase value
            #elif streams[i,0] in [45.0,21.0,22.0,23.0,34.0,37.0]: # saturated_water: 45, 21,22,23,34,37 
            elif streams[i,0] in sat_water: # saturated_water: 45, 21,22,23,34,37 
                streams[i,4] = 0.0    
            #elif streams[i,0] in [32.0,33.0,36.0,26.0,40.0]: # saturated_steam: 32,33,36, 26,40,
            elif streams[i,0] in sat_steam: # saturated_steam: 32,33,36, 26,40,
                streams[i,4] = 1.0
            else:
                streams[i,4] = -1.0

        #elif streams[i,0] in flue_gas:    # Flue gas stream
        else:      # For now, apply this to any other stream 
            streams[i,4] = -1.0

    #----------------------------------------------------------------------#
    
    # For gatex to read the file correctly, the
    # stream number should appear twice. Here we duplicate
    # the first column (which is the stream number)
    streams2 = insert(streams,1,streams[:,0],axis=1)
    head2    = insert(head,1,head[0])

    #---------------------------------------------------------------------##

    ## Create files for saving the different data parts ####################
    
    # Extract flow data from the stream array (with duplicate stream number!)
    # Then flatten to one vector and output to the flows.prn file
    flow_names = array(["STREAM","mdot","T","p","x"])
    inds = in1d(head2,flow_names)
    flows = streams2[:,inds]
    flows = flows.flatten()
    savetxt(gatex_f2,flows,fmt="%12.6f")
    flows2 = loadtxt(gatex_f2)    # To ensure file is written (Windows hack!)!
    print("Wrote file for GATEX: " + gatex_f2)

    # Now extract element composition data (with duplicate stream number!)
    # Then flatten to one vector and output to the composition.prn file
    element_names = array(["STREAM","Ar","CO2","CO","COS","H2O","CH4","H2","H2S","N2","O2","SO2"])
    inds = in1d(head2,element_names)
    elements = streams2[:,inds]
    elements = elements.flatten()
    savetxt(gatex_f3,elements,fmt="%12.6f")
    elements2 = loadtxt(gatex_f3)    # To ensure file is written (Windows hack!)!
    print("Wrote file for GATEX: " + gatex_f3)

    #---------------------------------------------------------------------##
    
    # Now call gatex #######################################################
    print("Calling GATEX...")

    # Determine whether to use wine or not
    # (If on linux or mac, use wine)
    try:
        uname = os.uname()    # Operating system info
        call_prefix = ""
        ###if ( uname[0] in ["Darwin"] ): call_prefix = "wine "
        if ( uname[0] in ["Darwin"] ): call_prefix = "/opt/local/bin/wine "
    except:
        call_prefix = ""
    
    ## Now call gatex and cross fingers !!! 
    os.system((call_prefix+gatex_exec))
    
    # If it worked, load the exergies
    E = loadtxt('exergies.m', skiprows = (37),comments = ']')
    
    # Add a first column that contains the stream number
    E = insert(E,0,streams[:,0],axis=1)

    # Also add an empty first row, so that the indices
    # match fontina's old code!
    E = insert(E,0,E[0,:]*0.0,axis=0)

    print('Checking GATEX output:')
    set_printoptions(precision=3,linewidth=150)
    print("Exergy table =")
    print("Columns: stream num., % m [kg/s], T [K], p[bar], H [MW], S [kW/K], EPH [MW], ECH [MW], E [MW]")
    print(E)    
    #---------------------------------------------------------------------##
    
    # Done, return the table of exergies
    # Dimensions: n_streams X n_vars
    return E


def exergy_to_excel(components,exergytable,filename="results.xlsx",sheetname='Results',newBook=True):
    '''
    Given the table of exergy results from GATEX, write them to
    an excel sheet.
    '''
    
    ## First, make a new empty workbook object if needed
    ## or load the existing workbook from the filename
    if newBook:
        
        # Generate a new workbook
        book = xl.Workbook()

        # Get the first sheet and rename it with the right name.
        sheet1 = book.get_active_sheet()    # Active sheet in new book is always the first one
        sheet1.title = sheetname            # Make sure the first sheet has the right name

    else:
        
        # Load the workbook from existing file
        book = xl.load_workbook(filename)
        
        # Get all existing sheetnames
        sheetnames = book.get_sheet_names()

        # If sheetname already exists, load it.
        # Otherwise, create a new sheet with the right name.
        if sheetname in sheetnames:
            sheet1 = book.get_sheet_by_name(sheetname)
        else:
            sheet1 = book.create_sheet()        # Create a new sheet
            sheet1.title = sheetname            # Make sure the new sheet has the right name
    
    # Make sure the results are input as a list
    # It should a list of individual components,
    # so if results is a scalar, then it must be
    # converted to a list containing one component
    if isscalar(components): components = [components]
    
    ncomps = len(components)
    
    # Get headings from component 1
    headings = array(components[0].keys())
    nhead  = len(headings)
    
    # Write all headings to sheet (first row)
    ###i = 1
    i = 2
    ###for j in arange(0,nhead):
    for j in arange(0,nhead):
        sheet1.cell(row=i,column=j+1).value = headings[j]
    
    # Loop over all components and write the row of variables
    for component in components:
        i = i + 1
        
        # Now loop over each variable and write it to the correct place
        for key, value in component.iteritems():

            # Which column for the current key?
            j = where(headings == key)
            if len(j)==1:
                j = j[0][0]   # Convert j from array into integer
                #print(i, j)
                sheet1.cell(row=i,column=j+1).value = value
    

    ## Now also write the complete exergy table to the file ###############
    
    # First generate some table header information
    #stream num., % m [kg/s], T [K], p[bar], H [MW], S [kW/K], EPH [MW], ECH [MW], E [MW]
    names = ['stream','m [kg/s]','T [K]','p [bar]','H [MW]',
             'S [kW/K]','EPH [MW]','ECH [MW]','E [MW]']
    
    # Write some information above the table location
    ###offset_row = ncomps + 12
    ###offset_col = 0
    offset_row = ncomps + 13
    offset_col = 1
    sheet1.cell(row=offset_row-3,column=offset_col).value = 'Exergy table'
    sheet1.cell(row=offset_row-2,column=offset_col).value = 'First row is zeros - ignore'

    ni = exergytable.shape[0]
    nj = exergytable.shape[1]
    
    # Write the table heading line first (by column)
    ###for j in arange(0,nj):
    for j in arange(1,nj):
        sheet1.cell(row=offset_row-1,column=offset_col+j).value = names[j]
    
    # Now loop over all table values and write to the sheet
    for j in arange(0,nj):
        for i in arange(0,ni):
            sheet1.cell(row=offset_row+i,column=offset_col+j).value = exergytable[i,j]

    # Save the book to the actual excel file
    book.save(filename)

    return
