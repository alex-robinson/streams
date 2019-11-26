
import os,sys,subprocess,string
from collections import OrderedDict
import openpyxl as xl                 # For writing/reading excel files

from copy import deepcopy
import re
import pkg_resources
from numpy import *

def load_reference(filename="ReferenceTables.xlsx",verbose=True):
    '''
    Load the reference data from the excel tables.
    '''

    # Load the workbook from existing file
    book = xl.load_workbook(filename)

    # Get all existing sheetnames
    #sheetnames = book.get_sheet_names()
    sheetnames = book

    # Load the reference table sheet - the first sheet
    sheet1 = book["SubstanceRefTables"]

    ## Now load some reference data

    # First get T0 and p0 used for these data
    T0 = sheet1['B2'].value
    p0 = sheet1['A2'].value

    # Now get the names of all the substances we
    # have data for (in column C) and the
    # relevant information

    refs = OrderedDict()

    for j in arange(3,1000):

        # First see if this row has a substance
        nm = sheet1.cell(row=j,column=3).value

        # If an element was actuall found in this row
        # store all the reference values in the OrderedDict!
        if not nm == None:
            refs[nm] = OrderedDict()
            refs[nm]['T0']      = T0
            refs[nm]['p0']      = p0
            refs[nm]['MW']      = sheet1.cell(row=j,column=4).value
            refs[nm]['e_ch_0a'] = sheet1.cell(row=j,column=5).value    # Ahrends
            refs[nm]['e_ch_0b'] = sheet1.cell(row=j,column=6).value    # Szargut
            refs[nm]['cp_0']    = sheet1.cell(row=j,column=8).value
            refs[nm]['h_0']     = sheet1.cell(row=j,column=9).value
            refs[nm]['s_0']     = sheet1.cell(row=j,column=10).value

            refs[nm]['H+']      = sheet1.cell(row=j,column=13).value
            refs[nm]['S+']      = sheet1.cell(row=j,column=14).value
            refs[nm]['a']       = sheet1.cell(row=j,column=15).value
            refs[nm]['b']       = sheet1.cell(row=j,column=16).value
            refs[nm]['c']       = sheet1.cell(row=j,column=17).value
            refs[nm]['d']       = sheet1.cell(row=j,column=18).value
        else:
            break

    # Add a default for unknown references
    nm = 'unknown'
    refs[nm] = deepcopy(refs['H2O'])
    refs[nm]['MW'] = 1.0
    refs[nm]['e_ch_0a'] = 0.0
    refs[nm]['e_ch_0b'] = 0.0
    refs[nm]['cp_0'] = 0.0
    refs[nm]['h_0']  = 0.0
    refs[nm]['s_0']  = 0.0
    refs[nm]['H+']   = None
    refs[nm]['S+']   = None
    refs[nm]['a']    = None
    refs[nm]['b']    = None
    refs[nm]['c']    = None
    refs[nm]['d']    = None


    text = "=" * 80
    text = text + "\nThe following substances are available for exergy calculations:\n"
    for i,key in enumerate(refs.keys()):
        text = text + key
        if not i+1 == len(refs.keys()): text = text + ", "
        if mod(i+1,12) == 0: text = text + "\n"
    text = text + "\n" + "=" * 80

    if verbose: print(text)

    return refs

### THE FUNCTIONS BELOW ARE METHODS
### FOR THE 'ref' object,
### which is an OrderedDict of OrderedDicts
### Eventually these should be encapsulated nicely!!!

def ref_check(refs,nm='N2'):
    '''
    Nicely print the reference values for the desired substance.
    '''
    print("Substance: " + nm)

    if nm in refs.keys():
        for key,value in refs[nm].items():
          print("    {:<10} = {}".format(key,value))

    else:
        print("Error: substance not found in reference values!")

    return

def ref_array(refs):

    # Get column/row headings
    tmp1,tmp2 = refs.items()[0]
    headings = tmp2.keys()
    names = refs.keys()

    # Determine length of output array
    ns = len(refs)
    nv = len(headings)

    # Generate array and fill it in
    table = zeros([ns,nv])
    i = -1
    for nm in names:
        i = i+1
        j = -1
        for key,val in refs[nm].items():
            j = j+1
            table[i,j] = val

    return(table)

def ref_print(refs):

    nms = refs
    for nm in refs.keys():
      print("Substance: " + nm)
      for key,value in refs[nm].items():
          print(" {:<10} = {}".format(key,value))

    return

### NOW STREAM CLASSES ###

class substance:
    '''
    This class manages calculations for a given substance.
    '''

    def __init__(self,id,name,T,p,mdot,x,T0=298.15,p0=1.013):

        # Define some constants
        R = 8.314       # kJ/kmol-K, ideal gas constant

        # Make a string of the id,name for easy formatting
        idstr = "{:4}".format(id)
        namestr = "{:11}".format(name)
        unknown = False

        # Adjust x values if they are tiny to simplify calcs.
        #if abs(x) < 1e-7: x = 0.0

        # Make sure x is valid
        if x < 0 or x > 1:
            err = '''
Stream {}: {}: Error: molar fraction x must be between 0 and 1!
'''.format(idstr,namestr)
            sys.exit(err)

        # Generate the initial state of this substance in the stream
        state = OrderedDict(T=T,p=p,mdot=mdot*x,x=x)

        # Load reference substance values
        REF_FILE = pkg_resources.resource_filename('streams', 'data/ReferenceTables.xlsx')
        refs     = load_reference(REF_FILE,verbose=False)

        # Now check that we can actually model this substance
        nameref = name
        if not name in refs.keys():
            if x > 0:
                print("Stream {}: {}: Warning: no reference values found for substance! " \
                      "Setting x = {:10.5g} => 0.0.".format(idstr,namestr,x))
            x = 0.0
            nameref = "unknown"
            unknown = True

        # Since we know the substance exists,
        # load up the reference values here
        ref = refs[nameref].copy()
        #ref = deepcopy(refs[nameref])  # Shallow copy should be good enough since we don't modify refs!

        # Check to make sure we are using the same reference state (T0,p0)
        if not (T0 == ref['T0'] and p0 == ref['p0']):
            err = '''
Stream {}: {}: Error: reference state is not the same as the reference state of the stream.
                      Stream (T0,p0) = ({:7.3f},{:7.3f})
                   Substance (T0,p0) = ({:7.3f},{:7.3f})
'''.format(idstr,namestr,T0,p0,ref['T0'],ref['p0'])
            sys.exit(err)

        # Get a useful factor (ajr: what is this?)
        y = T/1e3

        ##### Calculations of enthalpy h
        ##### (depends only on T, because we assume ideal gases)

        # If H+ not known for the element, then we
        # calculate it using cp_0, h_0 and T
        if ref['H+'] == None:
            state['h'] = ( (ref['cp_0']) * (T-ref['T0']) ) + ref['h_0']      # NO change in cp

        else:
            state['h'] = 1e3 * (  ref['H+'] + ref['a']*y
                                + ref['b']/2 * y**2
                                - ref['c']*y**(-1)
                                + ref['d']/3 * y**3 )

            # Limit enthalpy errors
            if state['h'] > -5.0 and state['h'] < 0.0:
                print("Stream {}: {}: Warning: small negative enthalpy set to zero.".format(idstr,namestr))
                state['h'] = 0.0

        ##### Calculations of entropy
        ##### (depend both on T and p)

        # If H+ not known for the element, then we
        # calculate it using cp_0, h_0 and T and p
        if ref['S+'] == None:
            state['s'] = ( (ref['cp_0']/T)*(T-ref['T0']) - R*log(p/ref['p0']) ) + ref['s_0']

        else:
            state['s'] = (  ref['S+'] + ref['a']*log(T) + ref['b']*y
                          - ref['c']/2 * y**(-2)
                          + ref['d']/2 * y**2   )     - R * log(p/ref['p0'])

        ##### Calculations of specific heat cp
        ##### (depends on T)
        if ref['a'] == None:
            state['cp'] = ref['cp_0']

        else:
            state['cp'] = ( ref['a'] + ref['b']*y
                          + ref['c'] * y**(-2)
                          + ref['d'] * y**2 )

        # Store the output
        self.id      = id
        self.idstr   = idstr
        self.name    = name
        self.namestr = namestr
        self.unknown = unknown
        self.state   = state
        self.ref     = ref

        return

    def __str__(self):
        '''Output the substance object to the screen in a human readable way.'''

        text = "\nStream {}: {}".format(self.idstr,self.namestr)
        for key,value in self.state.items():
            text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        text = text +  "\nREFERENCE VALUES"
        for key,value in self.ref.items():
            text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text


class stream:
    '''
    This class manages all values concerning a
    given stream and calculation of its exergy.
    '''

    def __init__(self,id,T,p,mdot,composition,phase=[1.0,0.0,0.0],T0=298.15,p0=1.013,exergy_type=None):
        '''
        Initialize a stream
          id = stream number/name
          T  = stream temperature
          p  = stream pressure
          mdot = stream total mass flow rate kg/hr ?
          composition = list of substances and molar fractions x
                        ( a list of tuples: [ ('N2',0.5), ('O2',0.1) ] )
          phase = [VFRAC,LFRAC,SFRAC], fraction of vapor, liquid and solid
                  *** phase only needed for H2O streams ***
          T0, p0 = reference state values for temp (K) and pressure (bar)
        '''

        ### TO DO ###
        # Implement optional arguments of an array row containing
        # all stream info and an array row containing the header
        # Then extract the info as needed for this class.
        #############

        # Make a string of the id for easy formatting
        idstr = "{:4}".format(id)

        # First store the state and the phase
        state = OrderedDict(T=T,p=p,mdot=mdot,T0=T0,p0=p0,phase=phase)

        ## Loop over the substances and initialize each
        ## one for the current stream state
        comp = OrderedDict()
        for name,x in composition:
            comp[name] = substance(id,name,T=T,p=p,mdot=mdot,x=x,T0=T0,p0=p0)
            #print(name + "\n")

        ## Calculate MW of the stream and
        ## the total molar fraction (should sum to 1 unless substances are missing)
        state['MW'] = 0.0
        xtot        = 0.0
        for key,sub in comp.items():
            state['MW'] = state['MW'] + ( sub.ref['MW']*sub.state['x'] )
            xtot        = xtot + sub.state['x']

        # Check if xtot is 1
        # if abs(1.0 - xtot) > 1e-5:
        #     print("Stream {}: Warning: total x not equal to 1!").format(idstr))

        ## Hack from Matlab: if a stream contains neither N2 or CH4, then set
        ## total weight fraction to water
        if (not 'N2' in comp.keys()) and (not 'CH4' in comp.keys()):
        #if comp['N2'].state['x'] + comp['CH4'].state['x'] == 0:
            comp['H2O'] = substance(id,'H2O',T=T,p=p,mdot=mdot,x=1.0,T0=T0,p0=p0)

        ## Handling water ##
        ## Make sure that if a stream has H2O(l), it also has H2O(g) and vice-versa
        ## (This will facilitate calculations later involving liquid and gas)
        if 'H2O(l)' in comp.keys() and not 'H2O' in comp.keys():
            comp['H2O'] = substance(id,'H2O',T=T,p=p,mdot=mdot,x=0.0,T0=T0,p0=p0)
        if 'H2O' in comp.keys() and not 'H2O(l)' in comp.keys():
            comp['H2O(l)'] = substance(id,'H2O(l)',T=T,p=p,mdot=mdot,x=0.0,T0=T0,p0=p0)

        ## Determine what kind of stream we have (H2O or not, flue gas or not)
        if 'H2O' in comp.keys():
            if comp['H2O'].state['x'] + comp['H2O(l)'].state['x'] == 1.0:
                isH2O = True
            else:
                isH2O = False
        else:
            isH2O = False

        #print("Stream {}: isH2O={}".format(idstr,isH2O))

        # Make a duplicate list of stream substances
        # that will be adjusted for calculating the standard (T0,p0) values
        comp0 = deepcopy(comp)

        ## If stream is flue gas, adjust water percentages
        ## (For now, always adjust since we don't know how to tell
        ##  if it's flue gas or not)
        # At 25C the pressure would be 0.0317bar, so:
        # x_H2O_new(g) = ( 0.0317 * ( 1- x_H2O ) ) / (1 - 0.0317)         # % mol
        # x_H2O(l) = x_H2O(l) + (x_H2O - x_H2O_new(g))
        if 'H2O(l)' in comp0.keys():
            x_l = comp0['H2O(l)'].state['x']
            x_g = comp0['H2O'].state['x']
            x       = x_l + x_g
            if x > 0.0:
                x_new_g = 0.0317*(xtot-x) / (xtot-0.0317)    # mol frac
                x_new_g = max(x_new_g,0.0)
                x_new_l = (x - x_new_g)                    # mol frac
                x_new_l = max(x_new_l,0.0)
                comp0['H2O(l)'].state['x'] = x_new_l
                comp0['H2O'].state['x']    = x_new_g

                text = (
"Stream {}: Check: x of H2O(l),H2O = {:5.3f},{:5.3f} => {:5.3f},{:5.3f}".format(idstr,x_l,x_g,x_new_l,x_new_g) )

                print(text)

            if not (comp0['H2O(l)'].state['x'] + comp0['H2O'].state['x']
                ==  comp['H2O(l)'].state['x'] +  comp['H2O'].state['x']):
                text = "Stream {}: Warning: corrected H2O molar fractions "\
                       "do not sum to original total!"\
                       " All H2O set to gas (x of H2O(l)=0).".format(idstr)
                print(text)
                comp0['H2O(l)'].state['x'] = 0.0
                comp0['H2O'].state['x']    = x

        ## Calculate enthalpy (h) and entropy (s) depending on type of stream
        if not isH2O:

            # Enthalpy and entropy
            state['h']    = 0.0    # Enthalpy at T  for mixture
            state['h_0']  = 0.0    # Enthalpy at T0 for mixture
            state['s']    = 0.0    # Entropy  at T  for mixture
            state['s_0']  = 0.0    # Entropy  at T0 for mixture

            # Loop over elements and get stream variables
            for key,sub in comp.items():
                state['h']   = state['h']   + ( sub.state['h']*sub.state['x'] )
                state['s']   = state['s']   + ( sub.state['s']*sub.state['x'] )

            # Loop over standard (T0,p0) elements and get stream variables
            for key,sub in comp0.items():
                state['h_0'] = state['h_0'] + ( sub.ref['h_0']*sub.state['x'] )
                state['s_0'] = state['s_0'] + ( sub.ref['s_0']*sub.state['x'] )


        else:   # Case 2: water stream, H2O(l) and/or H2O(g)

            #!!!! CHECK IF WE HAVE SATURATED WATER OR STEAM  !!!!!
            isSaturated = False
            ## ADD TEST HERE

            if isSaturated:
                # Calculate enthalpy from the IAPWS IF97 steam tables for saturated streams
                ## TO BE IMPLEMENTED !!!
                pass

            else:  # Not saturated, calculate here...

                # Copy enthalpy and entropy from the IAPWS IF97 steam tables for non-saturated streams
                # h_T_str  = ( (cp_T_el - cp_T0_el) * (T_str-T0) ) + h_T0_el        # here el: H2O
                # s_Tp_str = ( (cp_T_el - cp_T0_el) * ln(T_str/T0) + s_Tp0_el
                pass

            # For now, set values to zero, since we can't calculate anything
            state['h']   = 0.0
            state['s']   = 0.0
            state['h_0'] = 0.0
            state['s_0'] = 0.0
        ####

        # Make an additional output of H in MW and S in MW/K
        # H: kJ/kmol => MW; S: kJ/kmol/K => kW/K
        # H = h * mdot / ( (MW*1e-3) *1e6 )
        # S = s * mdot / ( (MW*1e-3) *1e3 )
        if not state['MW'] == 0.0:
            state['H'] = state['h']  * state['mdot'] / ((state['MW']*1e-3)*1e6)
            state['S'] = state['s']  * state['mdot'] / ((state['MW']*1e-3)*1e3)

        # Update stream information with locally defined values
        self.id    = id
        self.idstr = idstr
        self.isH2O = isH2O
        self.state = state
        self.comp  = comp
        self.comp0 = comp0

        if not exergy_type is None:
            # Finally, calculate exergy if desired

            if exergy_type == "gatex":
                GATEX_EXE_FILE = pkg_resources.resource_filename('streams', 'gatex_pc_if97_mj.exe')
                #print(GATEX_EXE_FILE)
                #sys.exit()
                self.calc_exergy_gatex(gatex_exec=GATEX_EXE_FILE)
            elif exergy_type in ["Ahrends","Sahrgut"]:
                self.calc_exergy(exergy_type)
            else:
                print("Stream {}: exergy_type not recognized: {}.".format(self.idstr,exergy_type))


        return

    def calc_exergy(self,exergy_type="Ahrends"):
        '''
        Calculation of exergies given the state of
        a stream.
        '''

        # Define some constants
        R = 8.314       # kJ/kmol-K, ideal gas constant

        # Get idstr
        idstr = self.idstr

        # Decide which chemical exergy value to use (Ahrends or Szargut)
        if exergy_type.lower() == "ahrends":
            name_ch = 'e_ch_0a'
        elif exergy_type.lower() == "szargut":
            name_ch = 'e_ch_0b'
        else:
            err = '''
Stream {}: Error: Incorrect exergy type given: {}
                    Only "ahrends" or "szargut" are allowed.
'''.format(idstr,exergy_type)
            sys.exit(err)


        # Open variables locally
        state = self.state
        comp  = self.comp
        comp0 = self.comp0

        ## PHYSICAL EXERGY
        # Physical specific exergy of stream  ###
        state['e_ph']    = ( state['h']-state['h_0']
                           - state['T0']*(state['s']-state['s_0']) )    # kJ/kmol

        # Limit e_ph to postive values and/or give warnings
        if state['e_ph'] < -5.0:
            print("Stream {}: Warning: negative physical exergy.".format(self.id))
        elif state['e_ph'] < 0.0:
            state['e_ph'] = 0.0


        ## CHEMICAL EXERGY
        # Chemical specific exergy of stream ###

        # If liquid water is present, e_ch must be calculated in two steps.
        # 1. Adjust the molar fractions to eliminate liquid water (so only gas phase is left)
        # 2. e_ch = e_ch_gas*x_tot_gas + e_ch_h2o(l)*x_tot_h2o(l)

        # Check for liquid water (at 25 C)
        if not self.isH2O and 'H2O(l)' in comp0.keys() \
                   and comp0['H2O(l)'].state['x'] > 0.0:

            # Make a temporary copy of the composition at 25 C
            comptmp = deepcopy(comp0)

            # Adjust molar fractions for liquid and gas calcs seperately

            # Get total molar fraction x_tot (should be 1, unless substances missing)
            x_tot = 0.0
            for key,sub in comp.items(): x_tot = x_tot + sub.state['x']

            # Determine the liquid and gas content
            x_tot_l = comptmp['H2O(l)'].state['x']
            x_tot_g = x_tot - x_tot_l

            # Get the scaling factor for the gases
            xfac = x_tot / x_tot_g

            ## Scale the x value and calculate the chemical exergy for gas
            x_tot_tmp = 0.0
            sum1 = 0.0
            sum2 = 0.0
            for key,sub in comptmp.items():
                if not key == "H2O(l)" and not sub.state['x'] == 0.0:
                    sub.state['x'] = sub.state['x']*xfac
                    x_tot_tmp = x_tot_tmp + sub.state['x']

                    sum1 = sum1 + ( sub.ref[name_ch]*sub.state['x'] )
                    sum2 = sum2 + ( sub.state['x']*log(sub.state['x']) )

                    #print("x {:12}: {}".format(key,sub.state['x']))

            e_ch_g = (sum1 + R*state['T0']*sum2) *x_tot_g           # kJ/kmol

            # print("x_tot     {}".format(x_tot))
            # print("x_tot_g   {}".format(x_tot_g))
            # print("x_tot_l   {}".format(x_tot_l))
            # print("x_tot_tmp {}".format(x_tot_tmp))

            # But now make sure that the temporary x_tot_tmp is the same as the original!
            if abs(x_tot-x_tot_tmp) > 1e-3:
                print("Stream {}: Error: x_tot_tmp not correct: {} != {}".format(idstr,x_tot_tmp,x_tot))
                sys.exit()

            ## Also get the chemical exergy for liquid water
            e_ch_l = comptmp['H2O(l)'].ref[name_ch]*comptmp['H2O(l)'].state['x']

            # Store total chemical exergy
            state['e_ch'] = e_ch_g + e_ch_l
        else:
            # No water present, calculate chemical exergy from composition

            sum1 = 0.0
            sum2 = 0.0
            for key,sub in comp.items():
                if not sub.state['x'] == 0.0:
                    sum1 = sum1 + ( sub.ref[name_ch]*sub.state['x'] )
                    sum2 = sum2 + ( sub.state['x']*log(sub.state['x']) )

            state['e_ch']    = sum1 + R*state['T0']*sum2           # [kJ/kmol]


        ## Convert exergy units [kJ/kmol] => [kJ/kg]
        state['e_ph_kg'] = state['e_ph'] /state['MW']          # [kJ/kg]
        state['e_ch_kg'] = state['e_ch'] /state['MW']          # [kJ/kg]

        ## Get absolute exergies [MW]
        state['E_ph'] = state['e_ph_kg'] *state['mdot']/1e3    # [MW]
        state['E_ch'] = state['e_ch_kg'] *state['mdot']/1e3    # [MW]

        ## TOTAL EXERGY

        # Specific
        state['e_tot_kg'] = state['e_ph_kg'] + state['e_ch_kg']   # [KJ/kg]

        # Absolute
        state['E_tot'] = state['E_ph'] + state['E_ch']            # [MW]

        E_tot_check = state['e_tot_kg'] * state['mdot'] /1e3      # [MW]

        ebs = 1e-5
        if not abs(state['E_tot']-E_tot_check) < ebs:
            err = '''
Stream {}: Warning: total exergy calculations do not match!
                  E_ph + E_ch != e_tot_kg * mdot
                  {:10.3f} + {:10.3f} != {:10.3f} * {:10.3f}
'''.format(idstr,state['E_ph'],state['E_ch'],state['e_tot_kg'],state['mdot']/1e3)

            #sys.exit(err)
            print(err)

        # Update the stream's state object
        self.state = state

        return

    def calc_exergy_gatex(self,fldr="./",gatex_exec="../gatex_pc_if97_mj.exe"):
        '''
        Use this subroutine to calculate exergy of the stream
        using GATEX.

        == INPUT ==
        stream  : stream object
        fldr    : folder containing gatex program and
                  where input and output data are stored
                  (for now must be the current working directory, since gatex.exe looks there)

        == OUTPUT ==
        state variables obtained from exergy calculations using GATEX:
        state['H','S','E_ph','E_ch','E_tot']
        '''

        # Define stream state variables locally
        state = self.state

        # Save stream information to gatex input files (gate.inp,flows.prn,composition.prn)
        self.stream_to_gatex_files()

        # Now call gatex #######################################################
        print("Calling GATEX: stream {:d}, T = {:6.1f}, P = {:6.3f}".format(self.id,state['T'],state['p']))

        # Determine whether to use wine or not
        # (If on linux or mac, use wine)
        try:
            uname = os.uname()    # Operating system info
            call_prefix = ""
            if ( uname[0] in ["Darwin"] ): call_prefix = "wine "
        except:
            call_prefix = ""

        ## Now call gatex and cross fingers !!!
        os.system((call_prefix+gatex_exec))
        #os.system((call_prefix+GATEX_EXE_FILE)

        # If it worked, load the exergies
        with open('exergies.m', 'rb') as f:
            contents = f.readlines()[37:]

            new = []
            for line in contents:
                a = unicode(line,"utf-8").strip().replace("\n","").replace("\r","").replace("*************","NaN").split()
                if not a[0] == "];": new.append(a)
                #print(a)
            E = asarray(new,dtype=float)

        # Add a first column that contains the stream number
        E = insert(E,0,self.id,axis=1)

        # # Also add an empty first row, so that the indices
        # # match fontina's old code!
        # E = insert(E,0,E[0,:]*0.0,axis=0)

        # print('Checking GATEX output:')
        # set_printoptions(precision=3,linewidth=150)
        # print("Exergy table =")
        # print("Columns: stream num., mdot [kg/s], T [K], p[bar], H [MW], S [kW/K], EPH [MW], ECH [MW], E [MW]")
        # print(E)
        #---------------------------------------------------------------------##

        ## Store the output values back in our stream object
        self.state['H']     = E[0,4]
        self.state['S']     = E[0,5]
        self.state['E_ph']  = E[0,6]
        self.state['E_ch']  = E[0,7]
        self.state['E_tot'] = E[0,8]

        # Done

        return

    def stream_to_gatex_files(self):
        '''Translate a stream object into an array readable by GATEX.'''

        ## CODE below to output in format for GATEX!!
        ## ajr: need to modify this so that we generate it from the streams object, not the data table.

        # Get local variable for current stream
        stream = self
        state  = stream.state

        # Make the output array and headings we want for each column
        # Note: here the heading 'x' stands for vfrac:vapor fraction; 'SE' stands for entropy
        header = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])

        try:
            id = float(stream.id)
        except:
            id = 1000+randint(1)
            print("Stream {}: Warning: generated random number for "\
                  "stream number to send to GATEX: {}".format(key,id))

        data    = zeros(len(header))
        data[0] = id
        data[1] = stream.state['mdot']
        data[2] = stream.state['T'] - 273.15
        data[3] = stream.state['p']
        data[4] = stream.state['phase'][0]    # vfrac is called 'x' in this table!
        data[5] = stream.state['s']           # Gatex doesn't actually use this apparently!

        # Now loop over substances in header
        # and insert the molar fraction into the table as necessary
        substances = stream.comp.keys()
        ioffset = 6
        for i,head in enumerate(header[ioffset:]):
            if head in substances:
                data[ioffset+i] = stream.comp[head].state['x']

        ## Eliminate nan's from the table, GATEX cannot read them
        data[isnan(data)] = 0.0

        # print('======================')
        # print("Generating GATEX input files")

        # Folder to save gatex files, hard coded to be the current directory,
        # since gatex will be called from here.
        fldr = "./"

        # Define important gatex filenames
        gatex_f1 = os.path.join(fldr,"gate.inp")
        gatex_f2 = os.path.join(fldr,"flows.prn")
        gatex_f3 = os.path.join(fldr,"composition.prn")

        # To help figure things out, create a header
        # that shows the variable of each column
        head = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])

        # How many streams and variables are we working with?
        n_streams = 1
        n_var     = len(head)

        ### Create the gate.inp file with the reference values #################
        text = '''

    [system]

    t0 = {t0:10.3f} ;
    p0 = {p0:10.3f} ;
    number = {ns:6d}. ;
    ndiff = {nd:6d}. ;
    kelvin = {K:4d}. ;

    '''.format(t0=state['T0']-273.15,p0=state['p0'],ns=n_streams,nd=n_streams,K=0)
        # print(text)

        #ref = open(gatex_f1,'w',buffering=0)
        ref = open(gatex_f1,'w')
        ref.write(text)
        ref.close()
        #print(gatex_f1 + " is closed? " + str(ref.closed))
        #print("Wrote file for GATEX: " + gatex_f1)

        #----------------------------------------------------------------------#

        # For gatex to read the file correctly, the
        # stream number should appear twice. Here we duplicate
        # the first column (which is the stream number)
        data2 = insert(data,1,data[0])
        head2 = insert(head,1,head[0])

        #---------------------------------------------------------------------##

        ## Create files for saving the different data parts ####################

        # Extract flow data from the stream array (with duplicate stream number!)
        # Then flatten to one vector and output to the flows.prn file
        flow_names = array(["STREAM","mdot","T","p","x"])
        inds = in1d(head2,flow_names)
        flows = data2[inds]
        flows = flows.flatten()
        savetxt(gatex_f2,flows,fmt="%10.4f")
        flows2 = loadtxt(gatex_f2)    # To ensure file is written (Windows hack!)!
        #print("Wrote file for GATEX: " + gatex_f2)

        # Now extract element composition data (with duplicate stream number!)
        # Then flatten to one vector and output to the composition.prn file
        element_names = array(["STREAM","Ar","CO2","CO","COS","H2O","CH4","H2","H2S","N2","O2","SO2"])
        inds = in1d(head2,element_names)
        elements = data2[inds]
        elements = elements.flatten()
        savetxt(gatex_f3,elements,fmt="%10.4f")
        elements2 = loadtxt(gatex_f3)    # To ensure file is written (Windows hack!)!
        #print("Wrote file for GATEX: " + gatex_f3)

        #---------------------------------------------------------------------##

        # Done! All three gatex input files have been written with stream information

        return

    def __repr__(self):

        text = "Stream: {}".format(self.id)

        for key,value in self.state.items():
            if key == "phase":
                text = text + "\n {:<12} = {:>5.2n},{:>5.2n},{:>5.2n}".format(key,float(value[0]),float(value[1]),float(value[2]))
            else:
                text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text

    def __str__(self):
        '''Output the stream object to the screen in a human readable way.'''

        text = "Stream: {}".format(self.id)

        for key,value in self.state.items():
            if key == "phase":
                text = text + "\n {:<12} = {:>5.2n},{:>5.2n},{:>5.2n}".format(key,float(value[0]),float(value[1]),float(value[2]))
            else:
                text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        # text = text +  "\n====== reference values ======"
        # for key,value in self.ref.items():
        #     text = text + "\n {:<12} = {:>12.7n}".format(key,float(value))

        return text


### SIMULATION CLASS ###
class simulation:
    '''This class holds all simulation information in the
       proper format. This can be read in from an excel file
       or original formats from aspen (not implemented) or ebsilon (not implemented)
       The simulation data is held in stream classes.
    '''

    def __init__(self,stream=None,filename="ExampleSimulation.xlsx",sheetname="ExampleStreams1",
                 exergy_method=None,exergy_type="Ahrends",
                 saturated_water=['-9999'],saturated_steam=['-9999'],flue_gas=['-9999']):
        '''Initialize the simulation'''

        if stream == None:
            ## Check file extension to see which type of data to load
            ext = filename.rsplit(".",1)[1]
            if ext == "xlsx":  # load excel
                self.streams = self.load_excel(filename=filename,sheetname=sheetname)
                newfilename = filename
            elif ext == "rep": # load aspen
                print(filename)
                self.streams = self.load_aspen(filename=filename)
                #newfilename = filename0.rsplit(".",1)[0] + ".xlsx"
            else:
                pass # Incorrect format or unimplemented format
        else:
            self.streams = OrderedDict()
            self.streams[stream.id] = stream

        ## Now we have the streams object


        # Adjust those streams that are 'saturated'
        # AJR: NOT READY YET!
        # saturated_water = [ '1', '2', '3' ]      # Stream numbers of saturated water
        # saturated_steam = [ '4', '5', '6' ]      # Stream numbers of saturated steam
        # flue_gas        = [ '7', '8' ]           # Stream numbers of flue gas

        ## NOTE: This should be inside of stream class!!!
        for key,stream in self.streams.items():

            # Total molar fraction of H2O in stream
            x_H2O_tot = stream.comp['H2O'].state['x'] + stream.comp['H2O(l)'].state['x']

            # Get the the current stream phase [g,l,s]
            phase = stream.state['phase']



            if x_H2O_tot == 1.0:            # H2O stream

                if phase[0] > 0.001 and phase[0] < 0.999:
                    pass  # Take original phase value
                elif stream.id in saturated_water:
                    phase = [0,1,0]
                elif stream.id in saturated_steam:
                    phase = [1,0,0]
                else:
                    phase = [-1,0,0]

            #elif stream.id in flue_gas:    # Flue gas stream
            else:      # For now, apply this to any other stream
                phase = [-1,0,0]

            stream.state['phase'] = phase

        # Output some summary information about the streams
        # (ie print all warnings, etc here)
        # TO DO!

        ## Calculate exergy now
        if exergy_method == "gatex":

            self.streams = self.calc_exergy_gatex(gatex_exec="../gatex_pc_if97_mj.exe")

        elif exergy_method == "default":

            for key,stream in self.streams.items():
                stream.calc_exergy(exergy_type=exergy_type)



        return

    def load_excel(self,filename="ExampleSimulation.xlsx",sheetname="ExampleStreams1"):
        '''Load the simulation data from an excel sheet.'''

        # Load the workbook from existing file
        book = xl.load_workbook(filename)

        # Get all existing sheetnames
        # sheetnames = book.get_sheet_names()
        sheetnames = book

        # Make sure the sheet exists!
        if not sheetname in sheetnames:
            err = '''
Error:: load_excel:: Desired sheetname {} does not exist in the workbook {}.
'''.format(filename,sheetname)
            sys.exit(err)

        # Load the simulation sheet
        sheet1 = book[sheetname]

        #### Begin loading simulation data

        streams = OrderedDict()

        for j in arange(2,1000):    # Max 1000 streams!

            # First see if this row has a substance
            streamid = sheet1.cell(row=j,column=1).value

            # If an element was actually found in this row,
            # load all the data and generate a new stream object
            if not streamid == None:

                mdot      = sheet1.cell(row=j,column=2).value
                T         = sheet1.cell(row=j,column=3).value
                p         = sheet1.cell(row=j,column=4).value
                # MW is in column 4, but is calculated internally
                vfrac     = sheet1.cell(row=j,column=6).value
                lfrac     = sheet1.cell(row=j,column=7).value
                sfrac     = sheet1.cell(row=j,column=8).value

                if vfrac == None: vfrac = 0.0
                if lfrac == None: lfrac = 0.0
                if sfrac == None: sfrac = 0.0

                # Now loop over substances
                comp = []
                for i in arange(0,500):   # Max 500 substances!

                    name = sheet1.cell(row=1,column=9+i).value

                    if not name == None:
                        name = name.strip()
                        x = sheet1.cell(row=j,column=9+i).value
                        if x == None: x = 0
                        #print(i,j,name,x)
                        comp.append( (name,x) )
                    else:
                        break

                print("Loading stream {}".format(streamid))

                # Now we have loaded the state variables and
                # the composition, so generate a stream object
                streamidstr = "{}".format(streamid)
                streams[streamidstr] = stream(id=streamid,T=T,p=p,mdot=mdot,
                                              phase=[vfrac,lfrac,sfrac],composition=comp)

            else:
                break

        return streams

    def write_excel(self,filename,sheetname,writeSim=True,writeResults=True,writeSubs=True):
        '''Write simulation results to an excel file.'''

        # Load streams locally
        streams = self.streams

        # Generate a sheetname for the results
        sheetname_results    = sheetname + "_exergy"
        sheetname_substances = sheetname + "_sub"

        # Check if file already exists
        isFile = os.path.isfile(filename)

        if isFile:
            # Open existing workbook
            book = xl.load_workbook(filename)

            # Check if our desired worksheet already exists
            # and adjust sheetname as needed.
            sheetnames = book.sheetnames
            if sheetname in sheetnames:
                print("{} already exists in {}.".format(sheetname,filename))
                choice = input("Overwrite these results (y/n)? ")
                if not choice in ["y","Y","yes","YES","Yes"]:
                    sheetname = sheetname + "_new"

            # If sheetname already exists, delete it so we can start fresh.
            if sheetname in sheetnames:
                sheet1 = book[sheetname]
                book.remove(sheet1)

            # Now generate a fresh sheet with the right name
            sheet1 = book.create_sheet()        # Create a new sheet
            sheet1.title = sheetname            # Make sure the new sheet has the right name

            # If sheetname already exists, delete it so we can start fresh.
            if sheetname_results in sheetnames:
                sheet2 = book[sheetname_results]
                book.remove(sheet2)

            # Make a second fresh sheet for the results
            sheet2 = book.create_sheet()
            sheet2.title = sheetname_results

            # If sheetname already exists, delete it so we can start fresh.
            if sheetname_substances in sheetnames:
                sheet3 = book[sheetname_substances]
                book.remove(sheet3)

            # Make a second fresh sheet for the results
            sheet3 = book.create_sheet()
            sheet3.title = sheetname_substances


        else:
            # Generate a new workbook
            book = xl.Workbook()

            # Get the first sheet and rename it with the right name.
            sheet1 = book.get_active_sheet()    # Active sheet in new book is always the first one
            sheet1.title = sheetname            # Make sure the first sheet has the right name

            # Make a second sheet for the results
            sheet2 = book.create_sheet()
            sheet2.title = sheetname_results

            # Make a third sheet for the substances
            sheet3 = book.create_sheet()
            sheet3.title = sheetname_substances

        ## Now we should have an open workbook
        ## with one or two empty sheets

        ## First write the simulation data to sheet1

        ## Make a header
        header1 = ["mdot","T","p",  "MW",    "VFRAC","LFRAC","SFRAC"]
        units1  = ["kg/s","K","bar","kg/mol","",     "",     ""     ]

        key = list(streams)[0]
        header2 = list(streams[key].comp)
        units2 = [""] * len(header2)

        header = header1 + header2
        units = units1 + units2

        # Define the column offset for the headers
        offset  = 2
        offset2 = 9

        sheet1.cell(row=2,column=1).value = "Stream"

        for j,head in enumerate(header):
            sheet1.cell(row=2,column=offset+j).value = head
            sheet1.cell(row=1,column=offset+j).value = units[j]

        # Loop over the streams and write each line of data to the file
        for i,key in enumerate(streams.keys()):

              state = streams[key].state

              # Write the stream id and state variables
              sheet1.cell(row=i+3,column=offset-1).value = key
              sheet1.cell(row=i+3,column=offset+0).value = state['mdot']
              sheet1.cell(row=i+3,column=offset+1).value = state['T']
              sheet1.cell(row=i+3,column=offset+2).value = state['p']
              sheet1.cell(row=i+3,column=offset+3).value = state['MW']
              sheet1.cell(row=i+3,column=offset+4).value = state['phase'][0]
              sheet1.cell(row=i+3,column=offset+5).value = state['phase'][1]
              sheet1.cell(row=i+3,column=offset+6).value = state['phase'][2]

              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header2):
                  if vnm in streams[key].comp.keys():
                      val = streams[key].comp[vnm]
                      sheet1.cell(row=i+3,column=offset2+j).value = val.state['x']

        ## Now write the results to sheet2

        ## Make a header
        #header = ["mdot","T","p","h","s","h_0","s_0","e_ph","e_ch","e_ph_kg","e_ch_kg","e_tot_kg","E_ph","E_ch","E_tot"]
        header = ["mdot","T","p","H","S","E_ph","E_ch","E_tot","h_0","s_0","h","s","e_ph","e_ch","e_ph_kg","e_ch_kg","e_tot_kg"]
        units  = ["kg/s","K","bar","MW","kW/K","MW","MW","MW","kJ/kmol","kJ/kmol/K","kJ/kmol","kJ/kmol/K","","","","",""]
        # Define the column offset for the header
        offset = 2

        sheet2.cell(row=2,column=1).value = "Stream"
        for j,head in enumerate(header):
            sheet2.cell(row=2,column=offset+j).value = head
            sheet2.cell(row=1,column=offset+j).value = units[j]

        # Loop over the streams and write each line of data to the file
        for i,key in enumerate(streams.keys()):

              # Write the stream id
              sheet2.cell(row=i+3,column=offset-1).value = key

              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header):
                  if vnm in streams[key].state.keys():
                      val = streams[key].state[vnm]
                      sheet2.cell(row=i+3,column=offset+j).value = val

        ## Now write the individual substance information in sheet3

        ## Make a header
        header1 = ["mdot","T","p","x","h","s","cp"]
        header2 = ["MW","T0","p0","e_ch_0a","e_ch_0b","cp_0","h_0","s_0","H+","S+","a","b","c","d"]
        header3 = ["x_25","h_0_25","s_0_25"]

        header  = header1 + header2 + header3

        col_header = "99CCFFFF"

        # Define the column offset for the header
        offset  = 4
        offset2 = len(header1) + offset
        offset3 = len(header1) + len(header2) + offset

        # Start a row counter
        row = 2

        # Loop over the streams and write each line of data to the file
        for i,key0 in enumerate(streams.keys()):

            row = row + 4

            # Write the header again (for each stream)
            sheet3.cell(row=row,column=2).value = "Stream"
            sheet3.cell(row=row,column=3).value = "Substance"
            for j,head in enumerate(header):
                sheet3.cell(row=row,column=offset+j).value = head

            # Cell background color
# _cell.style.fill.fill_type = Fill.FILL_SOLID
# _cell.style.fill.start_color.index = Color.DARKRED

            # Get the current stream
            stream = streams[key0]

            ## Now loop over each substance in the stream
            for j,key in enumerate(stream.comp):

              state  = stream.comp[key].state
              ref    = stream.comp[key].ref
              state0 = stream.comp0[key].state
              ref0   = stream.comp0[key].ref

              # Increment the row
              row = row + 2

              # Write the stream id and substance name
              sheet3.cell(row=row,column=offset-2).value = key0
              sheet3.cell(row=row,column=offset-1).value = key

              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header1):
                  val = state[vnm]
                  sheet3.cell(row=row,column=offset+j).value = val

              # Loop over each variable and
              # Write the variable to the sheet
              for j,vnm in enumerate(header2):
                  val = ref[vnm]
                  sheet3.cell(row=row,column=offset2+j).value = val

              # Output last ones manually since header doesn't match var names
              sheet3.cell(row=row,column=offset3+0).value = state0['x']
              sheet3.cell(row=row,column=offset3+1).value = ref0['h_0']
              sheet3.cell(row=row,column=offset3+2).value = ref0['s_0']

        ## Finally, save the book to the actual excel file
        book.save(filename)

        return

    def load_aspen(self,filename="simu1.rep"):
        '''
        Given the aspen output file located at fldr/file_in,
        convert the format to the stream format
        '''

        print("Read input from Aspen file: {}".format(filename))

        # Load lines of input file
        lines = open(filename,'r').readlines()

        # STEP 1: Filter just to all lines related to stream section

        # Names of subsections
        subsections = ["COMPONENTS: KMOL/HR","TOTAL FLOW:","STATE VARIABLES:"]
        nss = len(subsections)

        # Line substrings to skip entirely
        skipstrings = ["(CONTINUED)"]

        in_stream      = False
        in_subsection  = False
        got_subsection = [False] * nss
        skipnext       = False
        sectstart      = False

        # Define assumed units
        units_T = "C"
        units_P = "Bar"

        lines1 = []

        # Loop over the lines and add those inside of the 'stream section'
        # Modify some strings to make a proper table
        for line in lines:

            # First strip extra whitespace
            line = line.strip()
            addthisline = False

            ## Skip blank and other unwanted lines
            if line == "" or any([substring in line for substring in skipstrings]): continue

            ## Determine whether we are in the stream section
            if line == "STREAM SECTION":
                in_stream = True
                continue

            ## ... or not in the stream section anymore
            if "ASPEN PLUS" in line:
                in_stream = False
                continue

            # How many subsections have we looped over so far?
            ngss = sum(got_subsection)


            ## Determine if we're not in a subsection we want anymore
            if in_stream and in_subsection:
                if ":" in line: in_subsection = False

            ## And determine if now we found a subsection we want
            if in_stream and ngss < nss:
                if line == subsections[ngss]:
                    in_subsection        = True
                    got_subsection[ngss] = True
                    continue     # we don't actually want this line...

            ##
            ## Now that we know where we are (in_stream, in_subsection),
            ## Decide what to do with lines
            ##

            ## If we find a new stream heading, add this line to the list
            if in_stream and not in_subsection:

                if sectstart and "STREAM ID" in line:

                    # Make sure previous stream set was not empty
                    if len(lines1) > 1:
                        if "STREAM" in lines1[len(lines1)-1]: lines1.pop()

                    addthisline = True
                    sectstart = False

                # If the line only contains the character '-' it's a
                # section header...
                if all([c in "-" for c in line]):
                    sectstart = True
                    got_subsection = [False] *nss  # reset subsections

            ## ... Else if we're inside a subsection we want,
            ## then we'll add this line to the list
            elif in_stream and in_subsection:
                addthisline = True

            ## If we indeed found a line we want, then
            ## modify the strings and add it to the list!
            if addthisline:

                # Make some handy string replacements here
                line1 = line

                # Get some info about units if available
                if "TEMP   C"   in line1: units_T = "C"
                if "TEMP   K"   in line1: units_T = "K"
                if "PRES   BAR" in line1: units_p = "Bar"

                if "STREAM" in line1:
                    line1 = line1.replace("STREAM ID","STREAM   ")
                else:

                    # Remove extra units of T and p
                    line1 = line1.replace(" BAR ","     ")
                    line1 = line1.replace(" C ","   ")

                    # Change some headings
                    line1 = line1.replace("TEMP   C",  "T       ")   # Temperature
                    line1 = line1.replace("PRES   BAR","p         ") # Pressure
                    line1 = line1.replace("KG/HR","mdot ")           # Mass flow rate
                    line1 = line1.replace("KCAL/KG","H      ")       # Enthalpy
                    line1 = line1.replace("CAL/GM-K","SE      ")     # Entropy

                    ### Number corrections

                    # Replace missing numbers with nan
                    line1 = line1.replace("MISSING","nan")

                    # (exponents need the 'e', 1.0-05 => 1.0e-05)
                    matches = re.findall(r'[0-9]\-[0-9]',line1)
                    for substr in matches:
                        substrnew = substr[0:1] + "e-" + substr[2:3]
                        line1 = line1.replace(substr,substrnew)
                    matches = re.findall(r'[0-9]\+[0-9]',line1)
                    for substr in matches:
                        substrnew = substr[0:1] + "e" + substr[2:3]
                        line1 = line1.replace(substr,substrnew)


                # Now add this line to the subset of lines we want
                lines1.append(line1)
                #print(line1)

        # Remove last line if it belongs to a spurious heading...
        if "STREAM" in lines1[len(lines1)-1]:
            lines1.pop()

        ##########################

        # Split the lines into separate lists so that
        # they can properly be merged into an array
        tmplists = []
        q = -1
        for line in lines1:
            if "STREAM" in line:
                tmplists.append([])
                q = q + 1
            tmplists[q].append(line)

        # Check that all lists have the same length!!
        print("If these are all the same number, then data was properly read:")
        for q in arange(len(tmplists)):
            print("Length of each page:",len(tmplists[q]))

        # Save how man pages and how many columns (right now, rows)
        # of data are on each page.
        npages = len(tmplists)
        ncols  = len(tmplists[0])

        # Now, generate a new table, by pasting the multiples
        # on as additional columns. In this way, each column
        # represents one stream
        lines2 = tmplists[0]
        for q in arange(1,npages):
            for i in arange(0,ncols):
                lines2[i] = lines2[i] + "  " + tmplists[q][i]

        # Now split the lines by white space
        # Then loop through and remove columns
        # where the stream number is actually a string
        tmp = [line.split() for line in lines2]
        #headings = [val[0] for val in tmp]

        # Replace stream numbers that are strings with negative number
        for i in arange(len(tmp[0])):
            if not tmp[0][i] == "STREAM":
                try:
                    check = float(tmp[0][i])
                except:
                    tmp[0][i] = "-9999"

        # Split the lines by white space and convert list into
        # a numpy array. Transpose so that each row is one stream
        table0 = asarray( tmp )
        table0 = transpose(table0)

        # Extract the headings of each column and remove them from the data part
        headings = table0[0,:]

        # Modify some element names
        # Define aliases: (Aspen_name,Reference_name)
        aliases = [ ("WATER",  "H2O"),
                    ("H2O(g)", "H2O"),
                    (" AR ",     " Ar "),
                    (" S ",      " S(s) "),
                    (" C ",      " C(s) ") ]

        tmp =  "  "+"  ".join(headings)+"  "
        for alias in aliases:
            tmp = tmp.replace(alias[0],alias[1])
        headings = array(tmp.split())
        #print("  ".join(headings))

        inds = [ e in ["STREAM","-9999"] for e in table0[:,0] ]
        table1 = table0[logical_not(inds),:]

        # Convert the text table into a data array
        # and order the rows according to the stream number
        data = asarray( [[float(a) for a in row] for row in table1 ] )
        order = data[:,0].argsort()
        data = data[order,:]

        ### SO NOW ABOVE WE HAVE ALL THE SIMULATION DATA IN A TABLE
        ### Now, save it in the streams format

        # Indices of substances
        inds = [ not e in ['KMOL/HR','mdot','CUM/HR','TEMP','PRES','VFRAC','LFRAC','SFRAC']
                    for e in headings ]
        ii1 = find(inds)

        # Indices of other information
        inds = [ e in ['KMOL/HR','mdot','CUM/HR','TEMP','PRES','VFRAC','LFRAC','SFRAC']
                    for e in headings ]
        ii2 = find(inds)

        # Set all nan values to zero
        #data[isnan(data)] = 0

        # Unit converters
        conv_T = 0.0
        if units_T == "C": conv_T = 273.15
        conv_mdot = 1/3600.0

        ## Loop over each stream and load the data
        streams = OrderedDict()
        for j,row in enumerate(data):

            # Get state variables and id
            streamid = int(row[ii1[0]])
            tot      = row[ii2[0]]
            mdot     = row[ii2[1]] * conv_mdot # KG/HR => KG/S
            T        = row[ii2[3]] + conv_T    # Must be in Kelvin
            p        = row[ii2[4]]
            vfrac    = row[ii2[5]]
            lfrac    = row[ii2[6]]
            sfrac    = row[ii2[7]]

            # Check for reasonable temperature values.
            #print("Stream {}: T,p,mdot = {},{},{}".format(streamid,T,p,mdot))

            # Now loop over substances
            comp = []
            for i in ii1[1:]:   # All of ii1 indices, except the first that was the streamid

                name = headings[i]
                x    = row[i]
                if tot > 0.0: x = x / tot
                comp.append( (name,x) )

            # Store the variables inside a new stream object
            streamidstr = "{}".format(streamid)
            streams[streamidstr] = stream(id=streamid,T=T,p=p,mdot=mdot,
                                          phase=[vfrac,lfrac,sfrac],composition=comp)

        # Add to simulation object
        #self.streams = streams

        return streams

    def sim_to_gatex(self):
        '''Translate a simulation object into an array readable by GATEX.'''

        ## CODE below to output in format for GATEX!!
        ## ajr: need to modify this so that we generate it from the streams object, not the data table.

        # Define local variable for streams
        streams = self.streams

        # Get the shape of the data
        nstreams = len(streams)

        # Make the output array and headings we want for each column
        # Note: here the heading 'x' stands for vfrac:vapor fraction; 'SE' stands for entropy
        header = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])
        ncols = len(header)
        out = zeros((nstreams,ncols))

        # Loop over each stream (go row by row)
        for j,key in enumerate(streams.keys()):

            # Get local variable for current stream
            stream = streams[key]

            try:
                id = float(key)
            except:
                id = 1000+randint(1)
                print("Stream {}: Warning: generated random number for "\
                      "stream number to send to GATEX: {}".format(key,id))

            out[j,0] = id
            out[j,1] = stream.state['mdot']
            out[j,2] = stream.state['T']
            out[j,3] = stream.state['p']
            out[j,4] = stream.state['phase'][0]    # vfrac is called 'x' in this table!
            out[j,5] = stream.state['s']           # Gatex doesn't actually use this apparently!

            # Now loop over substances in header
            # and insert the molar fraction into the table as necessary
            substances = stream.comp.keys()
            ioffset = 6
            for i,head in enumerate(header[ioffset:]):
                if head in substances:
                    out[j,ioffset+i] = stream.comp[head].state['x']

        ## Eliminate nan's from the table, GATEX cannot read them
        out[isnan(out)] = 0.0

        # Reorder the data by stream number
        ### Sort the list by row (using the first column to decide the order) ##
        order = out[:,0].argsort()
        out = out[order,:]

        # Check that stream numbers go up sequentially,
        # if not, write a warning/error
        nsdiff = diff(out[:,0])
        inds = [not e == 1 for e in nsdiff]

        if any(inds):
            err ='''
ERROR:: sim_to_gatex:: Stream numbering error. "
  Each stream number should increase sequentially by 1."
  The following stream numbers appear to be out of sequence:"
'''
            for q in arange(len(out[:,0])-1):
                if inds[q]: print("    Stream ",int(out[q+1,0]))

            #sys.exit("Stream numbers (see above).")

        ## FINAL STEP: if desired, also write the data to output file in ebsilon format
        #if write_mfile: write_ebsilon(out,file_output)

        # print('Checking GATEX intput:')
        # set_printoptions(precision=3,linewidth=150)
        # print("Exergy table =")
        # print("Columns: stream num., mdot [kg/s], T [K], p[bar]")
        # print(out[:,0:3])

        # Return the array of stream data in ebsilon format
        return out

    def calc_exergy_gatex(self,fldr="./",gatex_exec="../gatex_pc_if97_mj.exe"):
        '''
        Use this subroutine to calculate exergy from a set of stream data (array)
        using GATEX.

        == INPUT ==
        streams : numpy array of dimensions n_streams X n_variables
        fldr    : folder containing gatex program and
                  where input and output data are stored
                  (for now must be the current working directory, since gatex.exe looks there)

        == OUTPUT ==
        E       : table of exergy calculations obtained from GATEX

        '''

        # First generate an array of stream info for GATEX
        data = self.sim_to_gatex()

        print('======================')
        print("Generating GATEX input files")

        # Define important gatex filenames
        gatex_f1 = os.path.join(fldr,"gate.inp")
        gatex_f2 = os.path.join(fldr,"flows.prn")
        gatex_f3 = os.path.join(fldr,"composition.prn")

        # To help figure things out, create a header
        # that shows the variable of each column
        head = array(["STREAM","mdot","T","p","x","SE","Ar","CO2","CO","COS","H2O","XX","CH4","H2","H2S","N2","O2","SO2","H"])

        # print("Temperatures:",data[:,2])
        # sys.exit()

        # How many streams and variables are we working with?
        n_streams = len(data)
        n_var     = len(head)

        ### Create the gate.inp file with the reference values #################
        text = '''

[system]

t0 = {t0:10.3f} ;
p0 = {p0:10.3f} ;
number = {ns:6d}. ;
ndiff = {nd:6d}. ;
kelvin = {K:4d}. ;

'''.format(t0=data[0,2],p0=data[0,3],ns=n_streams,nd=n_streams,K=1)
        # print(text)

        #ref = open(gatex_f1,'w',buffering=0)
        ref = open(gatex_f1,'w')
        ref.write(text)
        ref.close()
        print(gatex_f1 + " is closed? " + str(ref.closed))
        print("Wrote file for GATEX: " + gatex_f1)

        #----------------------------------------------------------------------#

        # For gatex to read the file correctly, the
        # stream number should appear twice. Here we duplicate
        # the first column (which is the stream number)
        data2 = insert(data,1,data[:,0],axis=1)
        head2    = insert(head,1,head[0])

        #---------------------------------------------------------------------##

        ## Create files for saving the different data parts ####################

        # Extract flow data from the stream array (with duplicate stream number!)
        # Then flatten to one vector and output to the flows.prn file
        flow_names = array(["STREAM","mdot","T","p","x"])
        inds = in1d(head2,flow_names)
        flows = data2[:,inds]
        flows = flows.flatten()
        savetxt(gatex_f2,flows,fmt="%10.4f")
        flows2 = loadtxt(gatex_f2)    # To ensure file is written (Windows hack!)!
        print("Wrote file for GATEX: " + gatex_f2)

        # Now extract element composition data (with duplicate stream number!)
        # Then flatten to one vector and output to the composition.prn file
        element_names = array(["STREAM","Ar","CO2","CO","COS","H2O","CH4","H2","H2S","N2","O2","SO2"])
        inds = in1d(head2,element_names)
        elements = data2[:,inds]
        elements = elements.flatten()
        savetxt(gatex_f3,elements,fmt="%10.4f")
        elements2 = loadtxt(gatex_f3)    # To ensure file is written (Windows hack!)!
        print("Wrote file for GATEX: " + gatex_f3)

        #---------------------------------------------------------------------##

        # Now call gatex #######################################################
        print("Calling GATEX...")

        # Determine whether to use wine or not
        # (If on linux or mac, use wine)
        try:
            uname = os.uname()    # Operating system info
            call_prefix = ""
            if ( uname[0] in ["Darwin"] ): call_prefix = "wine "
        except:
            call_prefix = ""

        ## Now call gatex and cross fingers !!!
        os.system((call_prefix+gatex_exec))


        # If it worked, load the exergies
        with open('exergies.m', 'rb') as f:
            contents = f.readlines()[37:]

            new = []
            for line in contents:
                a = unicode(line,"utf-8").strip().replace("\n","").replace("\r","").replace("*************","NaN").split()
                if not a[0] == "];": new.append(a)
                #print(a)
            E = asarray(new,dtype=float)

        # Add a first column that contains the stream number
        E = insert(E,0,data[:,0],axis=1)

        # # Also add an empty first row, so that the indices
        # # match fontina's old code!
        # E = insert(E,0,E[0,:]*0.0,axis=0)

        print('Checking GATEX output:')
        set_printoptions(precision=3,linewidth=150)
        print("Exergy table =")
        print("Columns: stream num., mdot [kg/s], T [K], p[bar], H [MW], S [kW/K], EPH [MW], ECH [MW], E [MW]")
        print(E)
        #---------------------------------------------------------------------##

        ## Store the output values back in our simulation object
        streams = self.streams
        keys = streams.keys()
        for j,idnum in enumerate(E[:,0]):
            idstr = str(int(idnum))
            #print("idstr: ",idstr)
            if idstr in keys:
                streams[idstr].state['H']     = E[j,4]
                streams[idstr].state['S']     = E[j,5]
                streams[idstr].state['E_ph']  = E[j,6]
                streams[idstr].state['E_ch']  = E[j,7]
                streams[idstr].state['E_tot'] = E[j,8]

        # Done

        return streams

    def __str__(self):
        '''Print out the handy simulation.'''

        text = ""

        for key,stream in self.streams.items():
            text = text + "\n" + repr(stream)


        return text
